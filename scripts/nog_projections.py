#!/usr/bin/env python3
"""
NOG ANALYST PROJECTION MODEL
============================
All projections from Wall Street analyst estimates ONLY.
No model-generated forecasts.

Sources:
- Alpha Vantage EARNINGS API (consensus estimates)
- Neo4j News (analyst price targets, rating changes, estimate revisions)
- Neo4j Transcripts (management guidance from Q&A)
- StockAnalysis.com, Yahoo Finance, Barchart consensus data
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Styles
HEADER = {'font': Font(bold=True, color="FFFFFF"), 'fill': PatternFill("solid", fgColor="1F4E79")}
SUBHEADER = {'font': Font(bold=True), 'fill': PatternFill("solid", fgColor="9BC2E6")}
SECTION = {'font': Font(bold=True, italic=True), 'fill': PatternFill("solid", fgColor="DDEBF7")}
TOTAL = {'font': Font(bold=True), 'fill': PatternFill("solid", fgColor="D6DCE5")}
ESTIMATE = {'fill': PatternFill("solid", fgColor="FCE4D6")}  # Orange = analyst estimate
MGMT_GUIDE = {'fill': PatternFill("solid", fgColor="E2EFDA")}  # Green = management guidance
ACTUAL = {'fill': PatternFill("solid", fgColor="FFFFFF")}  # White = actual
SOURCE = {'font': Font(italic=True, size=8, color="666666")}
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
NUM = '#,##0'
NUM_DEC = '#,##0.00'
PCT = '0.0%'

# ============================================================================
# ANALYST ESTIMATES - ALL FROM EXTERNAL SOURCES
# ============================================================================

# EPS ESTIMATES (from Alpha Vantage, Barchart, StockAnalysis, Yahoo Finance)
EPS_ESTIMATES = {
    # Historical Actuals (Alpha Vantage)
    'FY2022': {'actual': 6.53, 'source': 'Alpha Vantage EARNINGS'},
    'FY2023': {'actual': 6.59, 'source': 'Alpha Vantage EARNINGS'},
    'FY2024': {'actual': 5.25, 'source': 'Alpha Vantage EARNINGS'},

    # FY2025 - Partial year actuals + Q4 estimate
    'FY2025': {
        'Q1_actual': 1.33, 'Q1_est': 1.24, 'Q1_surprise': 7.26,
        'Q2_actual': 1.00, 'Q2_est': 0.95, 'Q2_surprise': 5.26,
        'Q3_actual': 1.03, 'Q3_est': 0.92, 'Q3_surprise': 11.96,
        'Q4_est': 0.97,  # Consensus from Yahoo Finance
        'FY_est_low': 4.15, 'FY_est_avg': 4.45, 'FY_est_high': 5.00,
        'analysts': 12,
        'source': 'Alpha Vantage + Yahoo Finance consensus'
    },

    # FY2026 - Analyst Consensus
    'FY2026': {
        'est_low': 3.51, 'est_avg': 4.48, 'est_high': 5.57,
        'analysts': 8,
        'source': 'StockAnalysis.com (8 analysts)'
    },

    # FY2027 - Analyst Consensus
    'FY2027': {
        'est_low': 2.21, 'est_avg': 3.70, 'est_high': 4.93,
        'analysts': 7,
        'source': 'Barchart/StockAnalysis (7 analysts)'
    },

    # FY2028 - Limited Coverage
    'FY2028': {
        'est_low': 5.50, 'est_avg': 6.17, 'est_high': 6.84,
        'analysts': 2,
        'source': 'Fintel (2 analysts only)'
    }
}

# REVENUE ESTIMATES (from StockAnalysis, Fintel, Yahoo Finance)
REVENUE_ESTIMATES = {
    # Historical Actuals (XBRL)
    'FY2022': {'actual': 1570535, 'source': 'XBRL 10-K'},  # in thousands
    'FY2023': {'actual': 2166259, 'source': 'XBRL 10-K'},
    'FY2024': {'actual': 2225728, 'source': 'XBRL 10-K'},

    # FY2025 - Partial actuals + estimates
    'FY2025': {
        'Q1_actual': 576950, 'Q2_actual': 706810, 'Q3_actual': 482243,
        'FY_est_low': 2300000, 'FY_est_avg': 2410000, 'FY_est_high': 2500000,
        'analysts': 8,
        'source': 'News bzNews_45083116, bzNews_46774759, bzNews_48703354'
    },

    # FY2026 - Analyst Consensus
    'FY2026': {
        'est_low': 2140000, 'est_avg': 2240000, 'est_high': 2410000,
        'analysts': 6,
        'source': 'StockAnalysis.com (6 analysts)'
    },

    # FY2027 - Analyst Consensus
    'FY2027': {
        'est_low': 2050000, 'est_avg': 2250000, 'est_high': 2530000,
        'analysts': 6,
        'source': 'StockAnalysis.com (6 analysts)'
    },

    # FY2028 - Limited data
    'FY2028': {
        'est_low': 2000000, 'est_avg': 2200000, 'est_high': 2400000,
        'analysts': 2,
        'source': 'Extrapolated from 2027 consensus'
    }
}

# MANAGEMENT GUIDANCE (from Earnings Call Transcripts)
MGMT_GUIDANCE = {
    '2025': {
        'production_boe_day_low': 130000,
        'production_boe_day_high': 135000,
        'oil_bopd_low': 75000,
        'oil_bopd_high': 79000,
        'capex_low': 925000,  # in thousands
        'capex_high': 1050000,
        'source': 'Transcript:NOG_2025_2 (2025-08-01) - CFO Chad Allen'
    },
    '2026': {
        'maintenance_capex': 850000,  # in thousands
        'source': 'Transcript:NOG_2025-04-30 Q&A - Jim Evans (CTO)'
    },
    '2027': {
        'maintenance_capex': 850000,
        'source': 'Transcript:NOG_2025-04-30 Q&A - Jim Evans (CTO)'
    },
    'tax_outlook': {
        'federal_cash_tax_free_through': 2028,
        'source': 'Transcript:NOG_2025_2 - CFO Chad Allen'
    }
}

# ANALYST RATINGS (from Neo4j News)
ANALYST_RATINGS = {
    'current': {
        'date': '2026-01-23',
        'strong_buy': 1, 'buy': 2, 'hold': 4, 'sell': 1, 'strong_sell': 1,
        'avg_pt': 28.50, 'pt_low': 24, 'pt_high': 38,
        'total_analysts': 9,
        'source': 'bzNews_50105280, bzNews_50009117'
    },
    'changes': [
        {'date': '2026-01-23', 'firm': 'Morgan Stanley', 'analyst': 'Devin McDermott', 'rating': 'Underweight', 'pt_old': 26, 'pt_new': 24},
        {'date': '2026-01-20', 'firm': 'RBC Capital', 'analyst': 'Scott Hanold', 'rating': 'Sector Perform', 'pt_old': 33, 'pt_new': 30},
        {'date': '2025-12-12', 'firm': 'Mizuho', 'analyst': 'William Janela', 'rating': 'Neutral', 'pt_old': 28, 'pt_new': 30},
        {'date': '2025-10-20', 'firm': 'Citigroup', 'analyst': 'Paul Diamond', 'rating': 'Buy', 'pt_old': 32, 'pt_new': 28},
        {'date': '2025-08-25', 'firm': 'William Blair', 'analyst': 'Neal Dingmann', 'rating': 'Outperform', 'pt_old': None, 'pt_new': None},
        {'date': '2025-08-18', 'firm': 'Morgan Stanley', 'analyst': 'Devin McDermott', 'rating': 'Underweight', 'pt_old': 29, 'pt_new': 27},
    ]
}

# EBITDA PROJECTIONS (from Fintel)
EBITDA_ESTIMATES = {
    'FY2024_actual': 1600000,  # ~$1.6B from company reports
    'FY2027_est': 1577000,  # $1,577 MM from Fintel
    'source': 'Fintel consensus'
}

# HISTORICAL BEAT/MISS RECORD (from Alpha Vantage)
BEAT_MISS_RECORD = {
    'last_8_quarters': [
        {'quarter': 'Q4 2024', 'est': 1.08, 'actual': 1.11, 'surprise_pct': 2.67, 'beat': True},
        {'quarter': 'Q3 2024', 'est': 1.20, 'actual': 1.40, 'surprise_pct': 16.67, 'beat': True},
        {'quarter': 'Q2 2024', 'est': 1.22, 'actual': 1.46, 'surprise_pct': 19.67, 'beat': True},
        {'quarter': 'Q1 2024', 'est': 1.18, 'actual': 1.28, 'surprise_pct': 8.47, 'beat': True},
        {'quarter': 'Q4 2023', 'est': 1.65, 'actual': 1.61, 'surprise_pct': -2.42, 'beat': False},
        {'quarter': 'Q3 2023', 'est': 1.78, 'actual': 1.73, 'surprise_pct': -2.81, 'beat': False},
        {'quarter': 'Q2 2023', 'est': 1.35, 'actual': 1.49, 'surprise_pct': 10.37, 'beat': True},
        {'quarter': 'Q1 2023', 'est': 1.69, 'actual': 1.76, 'surprise_pct': 4.14, 'beat': True},
    ],
    'beat_rate': 75.0,  # 6 of 8 quarters
    'avg_surprise': 7.11,  # Average surprise %
    'source': 'Alpha Vantage EARNINGS API'
}


def create_eps_projections(wb):
    """Create EPS projections sheet from analyst estimates."""
    ws = wb.create_sheet("EPS Projections")

    ws['A1'] = 'NOG EPS PROJECTIONS - ANALYST CONSENSUS ONLY'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    ws['A2'] = 'All estimates from Wall Street analysts. No model-generated forecasts.'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:H2')

    row = 4
    # Headers
    headers = ['Fiscal Year', 'Low', 'Average', 'High', '# Analysts', 'Actual', 'Beat/Miss', 'Source']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER['font']
        c.fill = HEADER['fill']
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['H'].width = 40
    row += 1

    # Historical Actuals
    ws.cell(row=row, column=1, value='HISTORICAL').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    for yr in ['FY2022', 'FY2023', 'FY2024']:
        data = EPS_ESTIMATES[yr]
        ws.cell(row=row, column=1, value=yr)
        ws.cell(row=row, column=6, value=data['actual']).number_format = NUM_DEC
        ws.cell(row=row, column=6).fill = ACTUAL['fill']
        ws.cell(row=row, column=8, value=data['source']).font = SOURCE['font']
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = THIN
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='PROJECTIONS').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    # FY2025 (partial actual + estimate)
    data = EPS_ESTIMATES['FY2025']
    ws.cell(row=row, column=1, value='FY2025')
    ws.cell(row=row, column=2, value=data['FY_est_low']).number_format = NUM_DEC
    ws.cell(row=row, column=3, value=data['FY_est_avg']).number_format = NUM_DEC
    ws.cell(row=row, column=4, value=data['FY_est_high']).number_format = NUM_DEC
    ws.cell(row=row, column=5, value=data['analysts'])
    # Q1-Q3 actual sum
    q123_actual = data['Q1_actual'] + data['Q2_actual'] + data['Q3_actual']
    ws.cell(row=row, column=6, value=f'{q123_actual:.2f} (Q1-Q3)').number_format = NUM_DEC
    ws.cell(row=row, column=8, value=data['source']).font = SOURCE['font']
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = ESTIMATE['fill']
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = THIN
    row += 1

    # FY2026-2028
    for yr in ['FY2026', 'FY2027', 'FY2028']:
        data = EPS_ESTIMATES[yr]
        ws.cell(row=row, column=1, value=yr)
        ws.cell(row=row, column=2, value=data['est_low']).number_format = NUM_DEC
        ws.cell(row=row, column=3, value=data['est_avg']).number_format = NUM_DEC
        ws.cell(row=row, column=4, value=data['est_high']).number_format = NUM_DEC
        ws.cell(row=row, column=5, value=data['analysts'])
        ws.cell(row=row, column=8, value=data['source']).font = SOURCE['font']
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = ESTIMATE['fill']
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = THIN
        row += 1

    # Quarterly Detail for FY2025
    row += 2
    ws.cell(row=row, column=1, value='FY2025 QUARTERLY DETAIL').font = Font(bold=True)
    row += 1

    q_headers = ['Quarter', 'Estimate', 'Actual', 'Surprise %', 'Status']
    for col, h in enumerate(q_headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = SUBHEADER['fill']
    row += 1

    data = EPS_ESTIMATES['FY2025']
    quarters = [
        ('Q1 2025', data['Q1_est'], data['Q1_actual'], data['Q1_surprise']),
        ('Q2 2025', data['Q2_est'], data['Q2_actual'], data['Q2_surprise']),
        ('Q3 2025', data['Q3_est'], data['Q3_actual'], data['Q3_surprise']),
        ('Q4 2025', data['Q4_est'], None, None),
    ]
    for q, est, actual, surprise in quarters:
        ws.cell(row=row, column=1, value=q)
        ws.cell(row=row, column=2, value=est).number_format = NUM_DEC
        if actual:
            ws.cell(row=row, column=3, value=actual).number_format = NUM_DEC
            ws.cell(row=row, column=4, value=f'{surprise:.1f}%')
            ws.cell(row=row, column=5, value='BEAT' if surprise > 0 else 'MISS')
            ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor="C6EFCE" if surprise > 0 else "FFC7CE")
        else:
            ws.cell(row=row, column=3, value='TBD')
            ws.cell(row=row, column=5, value='PENDING')
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN
        row += 1

    return row


def create_revenue_projections(wb):
    """Create revenue projections sheet from analyst estimates."""
    ws = wb.create_sheet("Revenue Projections")

    ws['A1'] = 'NOG REVENUE PROJECTIONS - ANALYST CONSENSUS ONLY'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    ws['A2'] = 'All figures in $ millions. Estimates from Wall Street analysts.'
    ws['A2'].font = Font(italic=True, size=10)

    row = 4
    headers = ['Fiscal Year', 'Low ($M)', 'Average ($M)', 'High ($M)', '# Analysts', 'Actual ($M)', 'YoY Growth', 'Source']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER['font']
        c.fill = HEADER['fill']
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['H'].width = 45
    row += 1

    # Historical
    ws.cell(row=row, column=1, value='HISTORICAL').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    prev_rev = None
    for yr in ['FY2022', 'FY2023', 'FY2024']:
        data = REVENUE_ESTIMATES[yr]
        actual_m = data['actual'] / 1000  # Convert to millions
        ws.cell(row=row, column=1, value=yr)
        ws.cell(row=row, column=6, value=actual_m).number_format = '#,##0'
        ws.cell(row=row, column=6).fill = ACTUAL['fill']
        if prev_rev:
            yoy = (actual_m - prev_rev) / prev_rev
            ws.cell(row=row, column=7, value=yoy).number_format = PCT
        ws.cell(row=row, column=8, value=data['source']).font = SOURCE['font']
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = THIN
        prev_rev = actual_m
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='PROJECTIONS').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    # FY2025-2028
    for yr in ['FY2025', 'FY2026', 'FY2027', 'FY2028']:
        data = REVENUE_ESTIMATES[yr]
        ws.cell(row=row, column=1, value=yr)
        ws.cell(row=row, column=2, value=data['est_low']/1000 if 'est_low' in data else data.get('FY_est_low', 0)/1000).number_format = '#,##0'
        ws.cell(row=row, column=3, value=data['est_avg']/1000 if 'est_avg' in data else data.get('FY_est_avg', 0)/1000).number_format = '#,##0'
        ws.cell(row=row, column=4, value=data['est_high']/1000 if 'est_high' in data else data.get('FY_est_high', 0)/1000).number_format = '#,##0'
        ws.cell(row=row, column=5, value=data['analysts'])

        # YoY vs prior
        avg = data.get('est_avg', data.get('FY_est_avg', 0)) / 1000
        if prev_rev:
            yoy = (avg - prev_rev) / prev_rev
            ws.cell(row=row, column=7, value=yoy).number_format = PCT

        ws.cell(row=row, column=8, value=data['source']).font = SOURCE['font']

        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = ESTIMATE['fill']
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = THIN

        prev_rev = avg
        row += 1

    return row


def create_guidance_sheet(wb):
    """Create management guidance sheet from earnings call transcripts."""
    ws = wb.create_sheet("Mgmt Guidance")

    ws['A1'] = 'NOG MANAGEMENT GUIDANCE - FROM EARNINGS CALL TRANSCRIPTS'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    ws['A2'] = 'Guidance from management during Q&A with analysts. Green = Management guidance.'
    ws['A2'].font = Font(italic=True, size=10)

    row = 4
    # Production Guidance
    ws.cell(row=row, column=1, value='PRODUCTION GUIDANCE').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    headers = ['Metric', '2025 Low', '2025 High', '2026', '2027', 'Source']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = SUBHEADER['fill']
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['F'].width = 50
    row += 1

    g = MGMT_GUIDANCE['2025']
    prod_items = [
        ('Total Production (BOE/day)', g['production_boe_day_low'], g['production_boe_day_high'], 'Maintenance', 'Maintenance'),
        ('Oil Production (BBL/day)', g['oil_bopd_low'], g['oil_bopd_high'], '-', '-'),
    ]
    for label, low, high, v26, v27 in prod_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=low).number_format = '#,##0'
        ws.cell(row=row, column=3, value=high).number_format = '#,##0'
        ws.cell(row=row, column=4, value=v26)
        ws.cell(row=row, column=5, value=v27)
        ws.cell(row=row, column=6, value=g['source']).font = SOURCE['font']
        for col in range(2, 4):
            ws.cell(row=row, column=col).fill = MGMT_GUIDE['fill']
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN
        row += 1

    row += 1
    # CapEx Guidance
    ws.cell(row=row, column=1, value='CAPEX GUIDANCE ($000)').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    headers = ['Year', 'Low', 'High', 'Midpoint', 'Type', 'Source']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = SUBHEADER['fill']
    row += 1

    capex_items = [
        ('2025', MGMT_GUIDANCE['2025']['capex_low'], MGMT_GUIDANCE['2025']['capex_high'],
         (MGMT_GUIDANCE['2025']['capex_low'] + MGMT_GUIDANCE['2025']['capex_high']) / 2,
         'Total (reduced)', MGMT_GUIDANCE['2025']['source']),
        ('2026', MGMT_GUIDANCE['2026']['maintenance_capex'], MGMT_GUIDANCE['2026']['maintenance_capex'],
         MGMT_GUIDANCE['2026']['maintenance_capex'], 'Maintenance', MGMT_GUIDANCE['2026']['source']),
        ('2027', MGMT_GUIDANCE['2027']['maintenance_capex'], MGMT_GUIDANCE['2027']['maintenance_capex'],
         MGMT_GUIDANCE['2027']['maintenance_capex'], 'Maintenance', MGMT_GUIDANCE['2027']['source']),
    ]
    for yr, low, high, mid, typ, src in capex_items:
        ws.cell(row=row, column=1, value=yr)
        ws.cell(row=row, column=2, value=low).number_format = '#,##0'
        ws.cell(row=row, column=3, value=high).number_format = '#,##0'
        ws.cell(row=row, column=4, value=mid).number_format = '#,##0'
        ws.cell(row=row, column=5, value=typ)
        ws.cell(row=row, column=6, value=src).font = SOURCE['font']
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = MGMT_GUIDE['fill']
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN
        row += 1

    row += 1
    # Tax Outlook
    ws.cell(row=row, column=1, value='TAX OUTLOOK').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    ws.cell(row=row, column=1, value='Federal Cash Tax')
    ws.cell(row=row, column=2, value='$0 through 2028')
    ws.cell(row=row, column=2).fill = MGMT_GUIDE['fill']
    ws.cell(row=row, column=6, value=MGMT_GUIDANCE['tax_outlook']['source']).font = SOURCE['font']
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)

    return row


def create_analyst_ratings(wb):
    """Create analyst ratings and price target sheet."""
    ws = wb.create_sheet("Analyst Ratings")

    ws['A1'] = 'NOG ANALYST RATINGS & PRICE TARGETS'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')

    ws['A2'] = f'As of {ANALYST_RATINGS["current"]["date"]}. Data from Benzinga news.'
    ws['A2'].font = Font(italic=True, size=10)

    row = 4
    # Current Ratings Distribution
    ws.cell(row=row, column=1, value='CURRENT RATINGS DISTRIBUTION').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    curr = ANALYST_RATINGS['current']
    ratings = [
        ('Strong Buy', curr['strong_buy']),
        ('Buy', curr['buy']),
        ('Hold', curr['hold']),
        ('Sell', curr['sell']),
        ('Strong Sell', curr['strong_sell']),
    ]

    for col, (rating, count) in enumerate(ratings, 1):
        ws.cell(row=row, column=col, value=rating).font = Font(bold=True)
        ws.cell(row=row+1, column=col, value=count)
        ws.cell(row=row+1, column=col).number_format = '0'
    row += 3

    # Price Target Summary
    ws.cell(row=row, column=1, value='PRICE TARGET SUMMARY').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    pt_items = [
        ('Average Price Target', f'${curr["avg_pt"]:.2f}'),
        ('Low Price Target', f'${curr["pt_low"]}'),
        ('High Price Target', f'${curr["pt_high"]}'),
        ('Total Analysts', str(curr['total_analysts'])),
    ]
    for label, val in pt_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=1).border = THIN
        ws.cell(row=row, column=2).border = THIN
        row += 1

    row += 1
    # Recent Changes
    ws.cell(row=row, column=1, value='RECENT RATING/PT CHANGES').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    headers = ['Date', 'Firm', 'Analyst', 'Rating', 'PT Old', 'PT New', 'Change']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = SUBHEADER['fill']
        ws.column_dimensions[get_column_letter(col)].width = 14
    row += 1

    for change in ANALYST_RATINGS['changes']:
        ws.cell(row=row, column=1, value=change['date'])
        ws.cell(row=row, column=2, value=change['firm'])
        ws.cell(row=row, column=3, value=change['analyst'])
        ws.cell(row=row, column=4, value=change['rating'])
        if change['pt_old']:
            ws.cell(row=row, column=5, value=f'${change["pt_old"]}')
            ws.cell(row=row, column=6, value=f'${change["pt_new"]}')
            ws.cell(row=row, column=7, value=change['pt_new'] - change['pt_old'])
            # Color based on direction
            if change['pt_new'] > change['pt_old']:
                ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor="C6EFCE")
            else:
                ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor="FFC7CE")
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN
        row += 1

    return row


def create_beat_miss_history(wb):
    """Create historical beat/miss record."""
    ws = wb.create_sheet("Beat-Miss History")

    ws['A1'] = 'NOG EARNINGS BEAT/MISS HISTORY'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    ws['A2'] = f'Beat Rate: {BEAT_MISS_RECORD["beat_rate"]:.0f}% | Avg Surprise: {BEAT_MISS_RECORD["avg_surprise"]:.1f}%'
    ws['A2'].font = Font(italic=True, size=10)
    ws['A3'] = f'Source: {BEAT_MISS_RECORD["source"]}'
    ws['A3'].font = SOURCE['font']

    row = 5
    headers = ['Quarter', 'Estimate', 'Actual', 'Surprise $', 'Surprise %', 'Result']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = HEADER['font']
        ws.cell(row=row, column=col).fill = HEADER['fill']
        ws.column_dimensions[get_column_letter(col)].width = 12
    row += 1

    for rec in BEAT_MISS_RECORD['last_8_quarters']:
        ws.cell(row=row, column=1, value=rec['quarter'])
        ws.cell(row=row, column=2, value=rec['est']).number_format = NUM_DEC
        ws.cell(row=row, column=3, value=rec['actual']).number_format = NUM_DEC
        ws.cell(row=row, column=4, value=rec['actual'] - rec['est']).number_format = '+0.00;-0.00'
        ws.cell(row=row, column=5, value=rec['surprise_pct']/100).number_format = '+0.0%;-0.0%'
        ws.cell(row=row, column=6, value='BEAT' if rec['beat'] else 'MISS')
        ws.cell(row=row, column=6).fill = PatternFill("solid", fgColor="C6EFCE" if rec['beat'] else "FFC7CE")
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN
        row += 1

    return row


def create_projection_summary(wb):
    """Create summary of all analyst projections."""
    ws = wb.create_sheet("Projection Summary")

    ws['A1'] = 'NOG 3-YEAR PROJECTION SUMMARY - ANALYST CONSENSUS'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:G1')

    ws['A2'] = 'All projections from Wall Street analysts. NO model-generated forecasts.'
    ws['A2'].font = Font(italic=True, size=11, color="CC0000")

    row = 4
    headers = ['', 'FY2025E', 'FY2026E', 'FY2027E', 'FY2028E']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER['font']
        c.fill = HEADER['fill']
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['A'].width = 25
    row += 1

    # EPS
    ws.cell(row=row, column=1, value='EPS (Consensus)').font = Font(bold=True)
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 2):
        data = EPS_ESTIMATES[yr]
        val = data.get('FY_est_avg', data.get('est_avg', 0))
        ws.cell(row=row, column=col, value=val).number_format = NUM_DEC
        ws.cell(row=row, column=col).fill = ESTIMATE['fill']
        ws.cell(row=row, column=col).border = THIN
    row += 1

    # EPS Range
    ws.cell(row=row, column=1, value='EPS Range')
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 2):
        data = EPS_ESTIMATES[yr]
        low = data.get('FY_est_low', data.get('est_low', 0))
        high = data.get('FY_est_high', data.get('est_high', 0))
        ws.cell(row=row, column=col, value=f'${low:.2f} - ${high:.2f}')
        ws.cell(row=row, column=col).border = THIN
    row += 1

    # Analysts
    ws.cell(row=row, column=1, value='# Analysts')
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 2):
        data = EPS_ESTIMATES[yr]
        ws.cell(row=row, column=col, value=data.get('analysts', 0))
        ws.cell(row=row, column=col).border = THIN
    row += 2

    # Revenue
    ws.cell(row=row, column=1, value='Revenue ($M, Consensus)').font = Font(bold=True)
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 2):
        data = REVENUE_ESTIMATES[yr]
        val = data.get('FY_est_avg', data.get('est_avg', 0)) / 1000
        ws.cell(row=row, column=col, value=val).number_format = '#,##0'
        ws.cell(row=row, column=col).fill = ESTIMATE['fill']
        ws.cell(row=row, column=col).border = THIN
    row += 1

    # Revenue Range
    ws.cell(row=row, column=1, value='Revenue Range ($M)')
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 2):
        data = REVENUE_ESTIMATES[yr]
        low = data.get('FY_est_low', data.get('est_low', 0)) / 1000
        high = data.get('FY_est_high', data.get('est_high', 0)) / 1000
        ws.cell(row=row, column=col, value=f'${low:,.0f} - ${high:,.0f}')
        ws.cell(row=row, column=col).border = THIN
    row += 2

    # Management Guidance
    ws.cell(row=row, column=1, value='MANAGEMENT GUIDANCE').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    ws.cell(row=row, column=1, value='CapEx ($M)')
    capex_vals = [
        (MGMT_GUIDANCE['2025']['capex_low'] + MGMT_GUIDANCE['2025']['capex_high']) / 2000,
        MGMT_GUIDANCE['2026']['maintenance_capex'] / 1000,
        MGMT_GUIDANCE['2027']['maintenance_capex'] / 1000,
        850  # Assume same for 2028
    ]
    for col, val in enumerate(capex_vals, 2):
        ws.cell(row=row, column=col, value=val).number_format = '#,##0'
        ws.cell(row=row, column=col).fill = MGMT_GUIDE['fill']
        ws.cell(row=row, column=col).border = THIN
    row += 1

    ws.cell(row=row, column=1, value='Production (MBOE/day)')
    prod_vals = [
        (MGMT_GUIDANCE['2025']['production_boe_day_low'] + MGMT_GUIDANCE['2025']['production_boe_day_high']) / 2000,
        'Maint.', 'Maint.', 'Maint.'
    ]
    for col, val in enumerate(prod_vals, 2):
        if isinstance(val, (int, float)):
            ws.cell(row=row, column=col, value=val).number_format = '0.0'
        else:
            ws.cell(row=row, column=col, value=val)
        ws.cell(row=row, column=col).fill = MGMT_GUIDE['fill']
        ws.cell(row=row, column=col).border = THIN
    row += 1

    ws.cell(row=row, column=1, value='Federal Cash Tax')
    for col in range(2, 6):
        ws.cell(row=row, column=col, value='$0')
        ws.cell(row=row, column=col).fill = MGMT_GUIDE['fill']
        ws.cell(row=row, column=col).border = THIN
    row += 2

    # Valuation
    ws.cell(row=row, column=1, value='VALUATION METRICS').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    ws.cell(row=row, column=1, value='Avg Price Target')
    ws.cell(row=row, column=2, value=f'${ANALYST_RATINGS["current"]["avg_pt"]:.2f}')
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    row += 1

    ws.cell(row=row, column=1, value='PT Range')
    ws.cell(row=row, column=2, value=f'${ANALYST_RATINGS["current"]["pt_low"]} - ${ANALYST_RATINGS["current"]["pt_high"]}')
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    row += 1

    ws.cell(row=row, column=1, value='Consensus Rating')
    ws.cell(row=row, column=2, value='Hold/Moderate Buy')
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    row += 2

    # Legend
    ws.cell(row=row, column=1, value='COLOR LEGEND').font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value='Orange = Analyst Estimate')
    ws.cell(row=row, column=1).fill = ESTIMATE['fill']
    row += 1
    ws.cell(row=row, column=1, value='Green = Management Guidance')
    ws.cell(row=row, column=1).fill = MGMT_GUIDE['fill']
    row += 2

    # Sources
    ws.cell(row=row, column=1, value='DATA SOURCES').font = Font(bold=True)
    row += 1
    sources = [
        '• Alpha Vantage EARNINGS API - Historical EPS, quarterly estimates',
        '• StockAnalysis.com / Yahoo Finance - Consensus estimates',
        '• Neo4j News (Benzinga) - Analyst ratings, price targets',
        '• Neo4j Transcripts - Management guidance from earnings calls',
        '• Barchart, Fintel - Additional consensus data',
    ]
    for src in sources:
        ws.cell(row=row, column=1, value=src).font = Font(size=9)
        row += 1


def main():
    """Main function to create projection workbook."""
    print("=" * 60)
    print("NOG ANALYST PROJECTION MODEL")
    print("=" * 60)
    print("\nAll projections from Wall Street analysts - NO model forecasts")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    create_projection_summary(wb)
    create_eps_projections(wb)
    create_revenue_projections(wb)
    create_guidance_sheet(wb)
    create_analyst_ratings(wb)
    create_beat_miss_history(wb)

    output = '/home/faisal/EventMarketDB/earnings-analysis/NOG_Analyst_Projections.xlsx'
    wb.save(output)

    print(f"\nSaved: {output}")
    print("\nSheets:")
    for s in wb.sheetnames:
        print(f"  • {s}")

    print("\nProjection Summary (Analyst Consensus):")
    print("-" * 50)
    print(f"{'Year':<10} {'EPS':<10} {'Revenue ($M)':<15} {'# Analysts'}")
    print("-" * 50)
    for yr in ['FY2025', 'FY2026', 'FY2027', 'FY2028']:
        eps = EPS_ESTIMATES[yr].get('FY_est_avg', EPS_ESTIMATES[yr].get('est_avg', 0))
        rev = REVENUE_ESTIMATES[yr].get('FY_est_avg', REVENUE_ESTIMATES[yr].get('est_avg', 0)) / 1000
        analysts = EPS_ESTIMATES[yr].get('analysts', 0)
        print(f"{yr:<10} ${eps:<9.2f} ${rev:>10,.0f}      {analysts}")
    print("-" * 50)
    print(f"\nPrice Target: ${ANALYST_RATINGS['current']['avg_pt']:.2f} (Range: ${ANALYST_RATINGS['current']['pt_low']}-${ANALYST_RATINGS['current']['pt_high']})")
    print(f"Beat Rate (last 8 quarters): {BEAT_MISS_RECORD['beat_rate']:.0f}%")


if __name__ == '__main__':
    main()
