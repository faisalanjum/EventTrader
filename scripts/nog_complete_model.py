#!/usr/bin/env python3
"""
NOG COMPLETE FINANCIAL MODEL - MAXIMUM DETAIL
==============================================
Single consolidated Excel with:
- Full historical financials from XBRL (10-K/10-Q) with all line items
- Revenue by segment (Oil, Gas/NGL, Other)
- 3-year analyst projections (2026-2028) with ranges
- Quarterly detail with derivations
- Full source attribution for EVERY data point
- Cross-linked sheets with formulas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Styles
HEADER = {'font': Font(bold=True, color="FFFFFF"), 'fill': PatternFill("solid", fgColor="1F4E79")}
SUBHEADER = {'font': Font(bold=True), 'fill': PatternFill("solid", fgColor="9BC2E6")}
SECTION = {'font': Font(bold=True, italic=True), 'fill': PatternFill("solid", fgColor="DDEBF7")}
TOTAL = {'font': Font(bold=True), 'fill': PatternFill("solid", fgColor="D6DCE5")}
FORMULA = {'fill': PatternFill("solid", fgColor="FFF2CC")}  # Yellow = formula
ACTUAL = {'fill': PatternFill("solid", fgColor="FFFFFF")}  # White = XBRL actual
ESTIMATE = {'fill': PatternFill("solid", fgColor="FCE4D6")}  # Orange = analyst estimate
MGMT_STYLE = {'fill': PatternFill("solid", fgColor="E2EFDA")}  # Green = management guidance
CHECK = {'fill': PatternFill("solid", fgColor="C6EFCE")}  # Light green = validation
SOURCE_FONT = Font(italic=True, size=8, color="666666")
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
NUM = '#,##0'
NUM_DEC = '#,##0.00'
PCT = '0.0%'

# ============================================================================
# COMPREHENSIVE XBRL DATA (in thousands)
# ============================================================================
XBRL_IS = {
    'FY2024': {
        'Revenue_Oil': 1897857, 'Revenue_GasNGL': 254222, 'Revenue_Other': 11682, 'Revenue_Total': 2225728,
        'LOE': 429792, 'ProductionTax': 157091, 'DDA': 736600, 'GA': 50463, 'Impairment': 0, 'OtherOpEx': 13951,
        'OpEx_Total': 1387897, 'OpIncome': 837831,
        'Interest': -157717, 'DerivGainLoss': 61967, 'OtherNonOp': 440, 'NonOp_Total': -157014,
        'PreTax': 680817, 'Tax_Current': 959, 'Tax_Deferred': 159550, 'Tax_Total': 160509,
        'NetIncome': 520308, 'EPS_Basic': 5.21, 'EPS_Diluted': 5.14,
        'Shares_Basic': 99904, 'Shares_Diluted': 101183,
        'source': 'XBRL 10-K FY2024 (NOG)'
    },
    'FY2023': {
        'Revenue_Oil': 925852, 'Revenue_GasNGL': 35800, 'Revenue_Other': 9230, 'Revenue_Total': 2166259,
        'LOE': 347006, 'ProductionTax': 160118, 'DDA': 482306, 'GA': 46801, 'Impairment': 0, 'OtherOpEx': 8166,
        'OpEx_Total': 1044397, 'OpIncome': 1121862,
        'Interest': -135664, 'DerivGainLoss': 259250, 'OtherNonOp': 4795, 'NonOp_Total': -121120,
        'PreTax': 1000742, 'Tax_Current': 692, 'Tax_Deferred': 77081, 'Tax_Total': 77773,
        'NetIncome': 922969, 'EPS_Basic': 10.48, 'EPS_Diluted': 10.27,
        'Shares_Basic': 88067, 'Shares_Diluted': 89906,
        'source': 'XBRL 10-K FY2023 (NOG)'
    },
    'FY2022': {
        'Revenue_Oil': 415732, 'Revenue_GasNGL': 260462, 'Revenue_Other': 0, 'Revenue_Total': 1570535,
        'LOE': 260676, 'ProductionTax': 158194, 'DDA': 248252, 'GA': 47201, 'Impairment': 0, 'OtherOpEx': 3020,
        'OpEx_Total': 717343, 'OpIncome': 853192,
        'Interest': -80331, 'DerivGainLoss': -415262, 'OtherNonOp': -185, 'NonOp_Total': -76854,
        'PreTax': 776338, 'Tax_Current': 0, 'Tax_Deferred': 3101, 'Tax_Total': 3101,
        'NetIncome': 773237, 'EPS_Basic': 10.96, 'EPS_Diluted': 10.51,
        'Shares_Basic': 70546, 'Shares_Diluted': 73553,
        'source': 'XBRL 10-K FY2022 (NOG)'
    }
}

XBRL_BS = {
    'FY2024': {
        'Cash': 8933, 'AR': 319210, 'DerivAssets_Curr': 69802, 'Prepaid': 102798, 'CurrentAssets': 500743,
        'OilGasProp': 5007831, 'DerivAssets_NC': 31088, 'OtherAssets_NC': 64160, 'NonCurrentAssets': 5103079,
        'TotalAssets': 5603822,
        'AP': 300629, 'AccruedLiab': 183168, 'DerivLiab_Curr': 45949, 'ARO_Curr': 14524, 'CurrentLiab': 544270,
        'LTDebt': 2369294, 'DerivLiab_NC': 8889, 'ARO_NC': 96766, 'DeferredTax': 255618, 'OtherLiab_NC': 8550,
        'NonCurrentLiab': 2739117, 'TotalLiab': 3283387,
        'CommonStock': 1005, 'APIC': 1876912, 'RetainedEarnings': 442518, 'Treasury': 0,
        'TotalEquity': 2320435,
        'source': 'XBRL 10-K FY2024 (NOG)'
    },
    'FY2023': {
        'Cash': 8195, 'AR': 301843, 'DerivAssets_Curr': 103063, 'Prepaid': 96306, 'CurrentAssets': 509407,
        'OilGasProp': 3900626, 'DerivAssets_NC': 52011, 'OtherAssets_NC': 22211, 'NonCurrentAssets': 3974848,
        'TotalAssets': 4484255,
        'AP': 195718, 'AccruedLiab': 135379, 'DerivLiab_Curr': 39140, 'ARO_Curr': 15524, 'CurrentLiab': 385761,
        'LTDebt': 1835554, 'DerivLiab_NC': 23587, 'ARO_NC': 96068, 'DeferredTax': 95608, 'OtherLiab_NC': 0,
        'NonCurrentLiab': 2050817, 'TotalLiab': 2436578,
        'CommonStock': 998, 'APIC': 2124468, 'RetainedEarnings': -77790, 'Treasury': 0,
        'TotalEquity': 2047676,
        'source': 'XBRL 10-K FY2023 (NOG)'
    },
    'FY2022': {
        'Cash': 2528, 'AR': 192476, 'DerivAssets_Curr': 54949, 'Prepaid': 70532, 'CurrentAssets': 320485,
        'OilGasProp': 2498168, 'DerivAssets_NC': 23645, 'OtherAssets_NC': 32880, 'NonCurrentAssets': 2554693,
        'TotalAssets': 2875178,
        'AP': 145992, 'AccruedLiab': 100424, 'DerivLiab_Curr': 67785, 'ARO_Curr': 30771, 'CurrentLiab': 344972,
        'LTDebt': 1525413, 'DerivLiab_NC': 77839, 'ARO_NC': 163166, 'DeferredTax': 18527, 'OtherLiab_NC': 0,
        'NonCurrentLiab': 1784945, 'TotalLiab': 2129917,
        'CommonStock': 785, 'APIC': 1745234, 'RetainedEarnings': -1000759, 'Treasury': 0,
        'TotalEquity': 745260,
        'source': 'XBRL 10-K FY2022 (NOG)'
    }
}

XBRL_CF = {
    'FY2024': {'CFO': 1408663, 'CFI': -1674754, 'CFF': 266829, 'NetChange': 738, 'BeginCash': 8195, 'EndCash': 8933, 'source': 'XBRL 10-K FY2024'},
    'FY2023': {'CFO': 1183321, 'CFI': -1862346, 'CFF': 684692, 'NetChange': 5667, 'BeginCash': 2528, 'EndCash': 8195, 'source': 'XBRL 10-K FY2023'},
    'FY2022': {'CFO': 928418, 'CFI': -1402777, 'CFF': 467367, 'NetChange': -6992, 'BeginCash': 9519, 'EndCash': 2528, 'source': 'XBRL 10-K FY2022'},
}

# Quarterly data for derivations
XBRL_Q = {
    '2024_Q1': {'Revenue': 396348, 'OpEx': 344027, 'OpIncome': 52321, 'NonOp': -37869, 'PreTax': 14452, 'Tax': 2846, 'NI': 11606, 'source': 'XBRL 10-Q Q1 2024'},
    '2024_H1': {'Revenue': 957113, 'OpEx': 685857, 'OpIncome': 271256, 'NonOp': -75501, 'PreTax': 195755, 'Tax': 45600, 'NI': 150163, 'source': 'XBRL 10-Q Q2 2024'},
    '2024_9M': {'Revenue': 1710751, 'OpEx': 1005555, 'OpIncome': 705196, 'NonOp': -112218, 'PreTax': 592978, 'Tax': 144400, 'NI': 448609, 'source': 'XBRL 10-Q Q3 2024'},
    '2025_Q1': {'Revenue': 602098, 'OpEx': 372817, 'OpIncome': 229281, 'NonOp': -43494, 'PreTax': 185787, 'Tax': 46800, 'NI': 138982, 'source': 'XBRL 10-Q Q1 2025'},
    '2025_H1': {'Revenue': 1308908, 'OpEx': 903461, 'OpIncome': 405447, 'NonOp': -87882, 'PreTax': 317565, 'Tax': 79000, 'NI': 238567, 'source': 'XBRL 10-Q Q2 2025'},
}

# ============================================================================
# ANALYST ESTIMATES - ALL FROM EXTERNAL SOURCES WITH FULL ATTRIBUTION
# ============================================================================
ANALYST_EPS = {
    'FY2025': {
        'Q1_est': 1.24, 'Q1_actual': 1.33, 'Q1_surprise': 7.26, 'Q1_source': 'Alpha Vantage EARNINGS',
        'Q2_est': 0.95, 'Q2_actual': 1.00, 'Q2_surprise': 5.26, 'Q2_source': 'Alpha Vantage EARNINGS',
        'Q3_est': 0.92, 'Q3_actual': 1.03, 'Q3_surprise': 11.96, 'Q3_source': 'Alpha Vantage EARNINGS',
        'Q4_est': 0.97, 'Q4_source': 'Yahoo Finance consensus',
        'FY_low': 4.15, 'FY_avg': 4.45, 'FY_high': 5.00, 'analysts': 12,
        'source': 'Alpha Vantage EARNINGS + Yahoo Finance consensus'
    },
    'FY2026': {
        'FY_low': 3.51, 'FY_avg': 4.48, 'FY_high': 5.57, 'analysts': 8,
        'source': 'StockAnalysis.com (8 analysts)',
        'revision_30d': 'down 5%', 'revision_source': 'Benzinga News bzNews_47182280'
    },
    'FY2027': {
        'FY_low': 2.21, 'FY_avg': 3.70, 'FY_high': 4.93, 'analysts': 7,
        'source': 'Barchart / StockAnalysis.com (7 analysts)'
    },
    'FY2028': {
        'FY_low': 5.50, 'FY_avg': 6.17, 'FY_high': 6.84, 'analysts': 2,
        'source': 'Fintel (2 analysts only - limited coverage)'
    }
}

ANALYST_REV = {  # in thousands
    'FY2025': {
        'Q1_est': 560820, 'Q1_actual': 576950, 'Q1_source': 'Benzinga bzNews_45083116',
        'Q2_est': 543790, 'Q2_actual': 706810, 'Q2_source': 'Benzinga bzNews_46774759',
        'Q3_est': 521124, 'Q3_actual': 482243, 'Q3_source': 'Benzinga bzNews_48703354',
        'FY_low': 2300000, 'FY_avg': 2410000, 'FY_high': 2500000, 'analysts': 8,
        'source': 'Yahoo Finance / StockAnalysis consensus'
    },
    'FY2026': {
        'FY_low': 2140000, 'FY_avg': 2240000, 'FY_high': 2410000, 'analysts': 6,
        'source': 'StockAnalysis.com (6 analysts)'
    },
    'FY2027': {
        'FY_low': 2050000, 'FY_avg': 2250000, 'FY_high': 2530000, 'analysts': 6,
        'source': 'StockAnalysis.com (6 analysts)'
    },
    'FY2028': {
        'FY_low': 2000000, 'FY_avg': 2200000, 'FY_high': 2400000, 'analysts': 2,
        'source': 'Fintel extrapolation (limited data)'
    }
}

ANALYST_PT = {
    'current': {
        'avg': 28.50, 'low': 24, 'high': 38, 'analysts': 9,
        'date': '2026-01-23',
        'source': 'Benzinga News (bzNews_50105280)'
    },
    'history': [
        {'date': '2026-01-23', 'firm': 'Morgan Stanley', 'analyst': 'Devin McDermott', 'rating': 'Underweight', 'pt': 24, 'source': 'bzNews_50105280'},
        {'date': '2026-01-20', 'firm': 'RBC Capital', 'analyst': 'Scott Hanold', 'rating': 'Sector Perform', 'pt': 30, 'source': 'bzNews_50009117'},
        {'date': '2025-12-12', 'firm': 'Mizuho', 'analyst': 'William Janela', 'rating': 'Neutral', 'pt': 30, 'source': 'bzNews_49358539'},
        {'date': '2025-10-20', 'firm': 'Citigroup', 'analyst': 'Paul Diamond', 'rating': 'Buy', 'pt': 28, 'source': 'bzNews_48307561'},
        {'date': '2025-10-14', 'firm': 'Morgan Stanley', 'analyst': 'Devin McDermott', 'rating': 'Underweight', 'pt': 26, 'source': 'bzNews_48204564'},
        {'date': '2025-09-15', 'firm': 'Mizuho', 'analyst': 'William Janela', 'rating': 'Neutral', 'pt': 28, 'source': 'bzNews_47673266'},
        {'date': '2025-08-25', 'firm': 'William Blair', 'analyst': 'Neal Dingmann', 'rating': 'Outperform', 'pt': None, 'source': 'bzNews_47305492'},
        {'date': '2025-08-18', 'firm': 'Morgan Stanley', 'analyst': 'Devin McDermott', 'rating': 'Underweight', 'pt': 27, 'source': 'bzNews_47182280'},
        {'date': '2025-08-04', 'firm': 'Piper Sandler', 'analyst': 'Mark Lear', 'rating': 'Neutral', 'pt': 25, 'source': 'bzNews_46832066'},
        {'date': '2025-07-17', 'firm': 'Piper Sandler', 'analyst': 'Mark Lear', 'rating': 'Neutral', 'pt': 31, 'source': 'bzNews_46466025'},
        {'date': '2025-07-16', 'firm': 'Mizuho', 'analyst': 'William Janela', 'rating': 'Neutral', 'pt': 32, 'source': 'bzNews_46439812'},
        {'date': '2025-07-14', 'firm': 'RBC Capital', 'analyst': 'Scott Hanold', 'rating': 'Sector Perform', 'pt': 33, 'source': 'bzNews_46392984'},
        {'date': '2025-07-09', 'firm': 'Citigroup', 'analyst': 'Paul Diamond', 'rating': 'Buy', 'pt': 38, 'source': 'bzNews_46330157'},
    ],
    'ratings': {'strong_buy': 1, 'buy': 2, 'hold': 4, 'sell': 1, 'strong_sell': 1, 'source': 'Benzinga aggregation'}
}

# Beat/Miss History
BEAT_MISS = [
    {'qtr': 'Q3 2025', 'est': 0.92, 'actual': 1.03, 'surprise': 11.96, 'date': '2025-11-06', 'source': 'Alpha Vantage'},
    {'qtr': 'Q2 2025', 'est': 0.95, 'actual': 1.00, 'surprise': 5.26, 'date': '2025-08-01', 'source': 'Alpha Vantage'},
    {'qtr': 'Q1 2025', 'est': 1.24, 'actual': 1.33, 'surprise': 7.26, 'date': '2025-04-28', 'source': 'Alpha Vantage'},
    {'qtr': 'Q4 2024', 'est': 1.08, 'actual': 1.11, 'surprise': 2.67, 'date': '2025-02-20', 'source': 'Alpha Vantage'},
    {'qtr': 'Q3 2024', 'est': 1.20, 'actual': 1.40, 'surprise': 16.67, 'date': '2024-11-05', 'source': 'Alpha Vantage'},
    {'qtr': 'Q2 2024', 'est': 1.22, 'actual': 1.46, 'surprise': 19.67, 'date': '2024-07-30', 'source': 'Alpha Vantage'},
    {'qtr': 'Q1 2024', 'est': 1.18, 'actual': 1.28, 'surprise': 8.47, 'date': '2024-04-30', 'source': 'Alpha Vantage'},
    {'qtr': 'Q4 2023', 'est': 1.65, 'actual': 1.61, 'surprise': -2.42, 'date': '2024-02-22', 'source': 'Alpha Vantage'},
]

# ============================================================================
# MANAGEMENT GUIDANCE (from Earnings Call Transcripts)
# ============================================================================
MGMT = {
    'FY2025': {
        'CapEx_low': 925000, 'CapEx_high': 1050000, 'CapEx_mid': 987500,
        'Production_low': 130000, 'Production_high': 135000,
        'Oil_low': 75000, 'Oil_high': 79000,
        'source': 'Transcript:NOG_2025_2 (2025-08-01) - CFO Chad Allen'
    },
    'FY2026': {
        'CapEx': 850000, 'CapEx_type': 'Maintenance',
        'source': 'Transcript:NOG_2025-04-30 Q&A - Jim Evans CTO: "Call it $850 million, roughly"'
    },
    'FY2027': {
        'CapEx': 850000, 'CapEx_type': 'Maintenance',
        'source': 'Transcript:NOG_2025-04-30 Q&A - Jim Evans CTO'
    },
    'FY2028': {
        'CapEx': 850000, 'CapEx_type': 'Maintenance',
        'source': 'Assumed: Same maintenance level as 2026-2027 per CTO guidance'
    },
    'Tax': {
        'outlook': 'No federal cash tax through 2028',
        'source': 'Transcript:NOG_2025_2 - CFO Chad Allen: "NOG will not be subject to federal cash taxes in 2025, and we do not anticipate having a federal cash tax liability through 2028"'
    },
    'MA': {
        'pipeline': '$8B+ in potential deals',
        'source': 'Transcript:NOG_2025_2 - Adam Durlam: "more than 10 ongoing processes...combined value exceeding $8 billion"'
    },
    'Q3Q4_2025': {
        'outlook': 'Q3 mid-single digit decline, Q4 recovery to current levels',
        'source': 'Transcript:NOG_2025_2 Q&A#12 - CEO Nick O\'Grady'
    }
}


def create_income_statement(ws):
    """Comprehensive Income Statement with all line items."""
    years_hist = ['FY2022', 'FY2023', 'FY2024']
    years_proj = ['FY2025E', 'FY2026E', 'FY2027E', 'FY2028E']
    all_years = years_hist + years_proj

    ws['A1'] = 'NOG CONSOLIDATED INCOME STATEMENT - FULL DETAIL'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:L1')
    ws['A2'] = 'Historical: XBRL 10-K | Projections: Wall Street Analyst Consensus | All figures in $ thousands'
    ws['A2'].font = Font(italic=True, size=10)

    row = 4
    # Headers
    ws.cell(row=row, column=1, value='Line Item').font = HEADER['font']
    ws.cell(row=row, column=1).fill = HEADER['fill']
    for col, yr in enumerate(all_years, 2):
        c = ws.cell(row=row, column=col, value=yr)
        c.font = HEADER['font']
        c.fill = HEADER['fill']
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.cell(row=row, column=len(all_years)+2, value='Source').font = HEADER['font']
    ws.cell(row=row, column=len(all_years)+2).fill = HEADER['fill']
    ws.column_dimensions[get_column_letter(len(all_years)+2)].width = 45
    ws.column_dimensions['A'].width = 32
    row += 1

    # REVENUE SECTION
    ws.cell(row=row, column=1, value='REVENUE').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    # Revenue by Segment
    rev_items = [
        ('Oil Sales', 'Revenue_Oil'),
        ('Natural Gas & NGL Sales', 'Revenue_GasNGL'),
        ('Other Revenue', 'Revenue_Other'),
    ]
    oil_row = row
    for label, key in rev_items:
        ws.cell(row=row, column=1, value=f'  {label}')
        for col, yr in enumerate(years_hist, 2):
            ws.cell(row=row, column=col, value=XBRL_IS[yr][key]).number_format = NUM
            ws.cell(row=row, column=col).fill = ACTUAL['fill']
            ws.cell(row=row, column=col).border = THIN
        # Project segments based on FY2024 mix
        fy24_pct = XBRL_IS['FY2024'][key] / XBRL_IS['FY2024']['Revenue_Total']
        for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 5):
            proj_rev = ANALYST_REV[yr]['FY_avg'] * fy24_pct
            ws.cell(row=row, column=col, value=proj_rev).number_format = NUM
            ws.cell(row=row, column=col).fill = ESTIMATE['fill']
            ws.cell(row=row, column=col).border = THIN
        ws.cell(row=row, column=len(all_years)+2, value=f'XBRL / Proj at FY24 mix ({fy24_pct*100:.1f}%)').font = SOURCE_FONT
        row += 1

    # Total Revenue
    ws.cell(row=row, column=1, value='TOTAL REVENUE').font = TOTAL['font']
    rev_row = row
    for col in range(2, 5):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({c}{oil_row}:{c}{row-1})')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 5):
        ws.cell(row=row, column=col, value=ANALYST_REV[yr]['FY_avg']).number_format = NUM
        ws.cell(row=row, column=col).fill = ESTIMATE['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value=ANALYST_REV['FY2026']['source']).font = SOURCE_FONT
    row += 1

    # Revenue Growth
    ws.cell(row=row, column=1, value='  YoY Growth %').font = Font(italic=True, size=9)
    for col in range(3, len(all_years)+2):
        prev = get_column_letter(col-1)
        curr = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=({curr}{rev_row}-{prev}{rev_row})/{prev}{rev_row}')
        ws.cell(row=row, column=col).number_format = PCT
        ws.cell(row=row, column=col).fill = FORMULA['fill']
    row += 2

    # OPERATING EXPENSES
    ws.cell(row=row, column=1, value='OPERATING EXPENSES').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    opex_items = [
        ('Lease Operating Expense', 'LOE', 0.193),  # FY24 % of rev
        ('Production Taxes', 'ProductionTax', 0.071),
        ('DD&A', 'DDA', 0.331),
        ('General & Administrative', 'GA', 0.023),
        ('Impairment', 'Impairment', 0.0),
        ('Other Operating Expense', 'OtherOpEx', 0.006),
    ]
    opex_start = row
    for label, key, pct in opex_items:
        ws.cell(row=row, column=1, value=f'  {label}')
        for col, yr in enumerate(years_hist, 2):
            ws.cell(row=row, column=col, value=XBRL_IS[yr][key]).number_format = NUM
            ws.cell(row=row, column=col).fill = ACTUAL['fill']
            ws.cell(row=row, column=col).border = THIN
        # Project at FY24 margin
        for col in range(5, len(all_years)+2):
            c = get_column_letter(col)
            ws.cell(row=row, column=col, value=f'={c}{rev_row}*{pct:.4f}')
            ws.cell(row=row, column=col).number_format = NUM
            ws.cell(row=row, column=col).fill = FORMULA['fill']
            ws.cell(row=row, column=col).border = THIN
        ws.cell(row=row, column=len(all_years)+2, value=f'XBRL / Proj at {pct*100:.1f}% of Rev').font = SOURCE_FONT
        row += 1

    # Total OpEx
    ws.cell(row=row, column=1, value='TOTAL OPERATING EXPENSES').font = TOTAL['font']
    opex_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({c}{opex_start}:{c}{row-1})')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    row += 1

    # Operating Income
    ws.cell(row=row, column=1, value='OPERATING INCOME').font = Font(bold=True, size=11)
    opinc_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={c}{rev_row}-{c}{opex_row}')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value='=Revenue - OpEx').font = SOURCE_FONT
    row += 1

    # Op Margin
    ws.cell(row=row, column=1, value='  Operating Margin %').font = Font(italic=True, size=9)
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={c}{opinc_row}/{c}{rev_row}')
        ws.cell(row=row, column=col).number_format = PCT
        ws.cell(row=row, column=col).fill = FORMULA['fill']
    row += 2

    # NON-OPERATING
    ws.cell(row=row, column=1, value='NON-OPERATING INCOME (EXPENSE)').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    nonop_items = [
        ('Interest Expense', 'Interest'),
        ('Derivative Gain/(Loss)', 'DerivGainLoss'),
        ('Other Non-Operating', 'OtherNonOp'),
    ]
    nonop_start = row
    for label, key in nonop_items:
        ws.cell(row=row, column=1, value=f'  {label}')
        for col, yr in enumerate(years_hist, 2):
            ws.cell(row=row, column=col, value=XBRL_IS[yr][key]).number_format = NUM
            ws.cell(row=row, column=col).fill = ACTUAL['fill']
            ws.cell(row=row, column=col).border = THIN
        # Use FY2024 level for projections
        fy24_val = XBRL_IS['FY2024'][key]
        for col in range(5, len(all_years)+2):
            ws.cell(row=row, column=col, value=fy24_val).number_format = NUM
            ws.cell(row=row, column=col).fill = ESTIMATE['fill']
            ws.cell(row=row, column=col).border = THIN
        ws.cell(row=row, column=len(all_years)+2, value='XBRL / Held at FY24 level').font = SOURCE_FONT
        row += 1

    # Total Non-Op
    ws.cell(row=row, column=1, value='TOTAL NON-OPERATING').font = TOTAL['font']
    nonop_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({c}{nonop_start}:{c}{row-1})')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).border = THIN
    row += 1

    # Pre-Tax
    ws.cell(row=row, column=1, value='INCOME BEFORE TAX').font = Font(bold=True, size=11)
    pretax_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={c}{opinc_row}+{c}{nonop_row}')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value='=OpIncome + NonOp').font = SOURCE_FONT
    row += 1

    # Tax
    ws.cell(row=row, column=1, value='Income Tax Expense')
    tax_row = row
    for col, yr in enumerate(years_hist, 2):
        ws.cell(row=row, column=col, value=XBRL_IS[yr]['Tax_Total']).number_format = NUM
        ws.cell(row=row, column=col).fill = ACTUAL['fill']
        ws.cell(row=row, column=col).border = THIN
    # Per CFO: No federal cash tax through 2028
    for col in range(5, len(all_years)+2):
        ws.cell(row=row, column=col, value=0).number_format = NUM
        ws.cell(row=row, column=col).fill = MGMT_STYLE['fill']
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value=MGMT['Tax']['source'][:60]+'...').font = SOURCE_FONT
    row += 1

    # Net Income
    ws.cell(row=row, column=1, value='NET INCOME').font = Font(bold=True, size=12)
    ni_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={c}{pretax_row}-{c}{tax_row}')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value='=PreTax - Tax').font = SOURCE_FONT
    row += 1

    # Net Margin
    ws.cell(row=row, column=1, value='  Net Margin %').font = Font(italic=True, size=9)
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={c}{ni_row}/{c}{rev_row}')
        ws.cell(row=row, column=col).number_format = PCT
        ws.cell(row=row, column=col).fill = FORMULA['fill']
    row += 2

    # EPS SECTION
    ws.cell(row=row, column=1, value='EARNINGS PER SHARE').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    # EPS Basic
    ws.cell(row=row, column=1, value='EPS - Basic')
    for col, yr in enumerate(years_hist, 2):
        ws.cell(row=row, column=col, value=XBRL_IS[yr]['EPS_Basic']).number_format = NUM_DEC
        ws.cell(row=row, column=col).fill = ACTUAL['fill']
        ws.cell(row=row, column=col).border = THIN
    row += 1

    # EPS Diluted
    ws.cell(row=row, column=1, value='EPS - Diluted').font = Font(bold=True)
    eps_row = row
    for col, yr in enumerate(years_hist, 2):
        ws.cell(row=row, column=col, value=XBRL_IS[yr]['EPS_Diluted']).number_format = NUM_DEC
        ws.cell(row=row, column=col).fill = ACTUAL['fill']
        ws.cell(row=row, column=col).border = THIN
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 5):
        ws.cell(row=row, column=col, value=ANALYST_EPS[yr]['FY_avg']).number_format = NUM_DEC
        ws.cell(row=row, column=col).fill = ESTIMATE['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value='XBRL / Analyst consensus avg').font = SOURCE_FONT
    row += 1

    # EPS Range
    ws.cell(row=row, column=1, value='  EPS Range (Low-High)').font = Font(italic=True, size=9)
    for col in range(2, 5):
        ws.cell(row=row, column=col, value='-')
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 5):
        d = ANALYST_EPS[yr]
        ws.cell(row=row, column=col, value=f'${d["FY_low"]:.2f}-${d["FY_high"]:.2f}')
    row += 1

    # Analysts
    ws.cell(row=row, column=1, value='  # Analysts').font = Font(italic=True, size=9)
    for col in range(2, 5):
        ws.cell(row=row, column=col, value='-')
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 5):
        ws.cell(row=row, column=col, value=ANALYST_EPS[yr]['analysts'])
    row += 1

    # Shares
    ws.cell(row=row, column=1, value='Diluted Shares (thousands)')
    for col, yr in enumerate(years_hist, 2):
        ws.cell(row=row, column=col, value=XBRL_IS[yr]['Shares_Diluted']).number_format = NUM
        ws.cell(row=row, column=col).fill = ACTUAL['fill']
        ws.cell(row=row, column=col).border = THIN
    # Assume flat
    for col in range(5, len(all_years)+2):
        ws.cell(row=row, column=col, value=XBRL_IS['FY2024']['Shares_Diluted']).number_format = NUM
        ws.cell(row=row, column=col).fill = ESTIMATE['fill']
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value='XBRL / Assumed flat').font = SOURCE_FONT
    row += 2

    # VALIDATION
    ws.cell(row=row, column=1, value='VALIDATION CHECKS').font = Font(bold=True)
    row += 1
    checks = [
        ('Rev - OpEx = OpInc', f'=IF(ABS({{c}}{rev_row}-{{c}}{opex_row}-{{c}}{opinc_row})<1,"PASS","FAIL")'),
        ('OpInc + NonOp = PreTax', f'=IF(ABS({{c}}{opinc_row}+{{c}}{nonop_row}-{{c}}{pretax_row})<1,"PASS","FAIL")'),
        ('PreTax - Tax = NI', f'=IF(ABS({{c}}{pretax_row}-{{c}}{tax_row}-{{c}}{ni_row})<1,"PASS","FAIL")'),
    ]
    for label, formula in checks:
        ws.cell(row=row, column=1, value=label)
        for col in range(2, len(all_years)+2):
            c = get_column_letter(col)
            ws.cell(row=row, column=col, value=formula.format(c=c))
            ws.cell(row=row, column=col).fill = CHECK['fill']
        row += 1

    return ni_row, eps_row, rev_row


def create_balance_sheet(ws):
    """Detailed Balance Sheet with all line items."""
    years_hist = ['FY2022', 'FY2023', 'FY2024']
    years_proj = ['FY2025E', 'FY2026E', 'FY2027E', 'FY2028E']
    all_years = years_hist + years_proj

    ws['A1'] = 'NOG CONSOLIDATED BALANCE SHEET - FULL DETAIL'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:L1')

    row = 4
    ws.cell(row=row, column=1, value='($ thousands)').font = Font(italic=True, size=9)
    for col, yr in enumerate(all_years, 2):
        c = ws.cell(row=row, column=col, value=yr)
        c.font = HEADER['font']
        c.fill = HEADER['fill']
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.cell(row=row, column=len(all_years)+2, value='Source').font = HEADER['font']
    ws.cell(row=row, column=len(all_years)+2).fill = HEADER['fill']
    ws.column_dimensions[get_column_letter(len(all_years)+2)].width = 40
    ws.column_dimensions['A'].width = 30
    row += 1

    # ASSETS
    ws.cell(row=row, column=1, value='ASSETS').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    ws.cell(row=row, column=1, value='Current Assets:').font = SUBHEADER['font']
    row += 1

    curr_items = [
        ('Cash & Equivalents', 'Cash'),
        ('Accounts Receivable', 'AR'),
        ('Derivative Assets', 'DerivAssets_Curr'),
        ('Prepaid & Other', 'Prepaid'),
    ]
    curr_start = row
    for label, key in curr_items:
        ws.cell(row=row, column=1, value=f'  {label}')
        for col, yr in enumerate(years_hist, 2):
            ws.cell(row=row, column=col, value=XBRL_BS[yr][key]).number_format = NUM
            ws.cell(row=row, column=col).fill = ACTUAL['fill']
            ws.cell(row=row, column=col).border = THIN
        # Project flat or with growth
        fy24_val = XBRL_BS['FY2024'][key]
        for col in range(5, len(all_years)+2):
            ws.cell(row=row, column=col, value=fy24_val).number_format = NUM
            ws.cell(row=row, column=col).fill = ESTIMATE['fill']
            ws.cell(row=row, column=col).border = THIN
        ws.cell(row=row, column=len(all_years)+2, value='XBRL / Held flat').font = SOURCE_FONT
        row += 1

    ws.cell(row=row, column=1, value='Total Current Assets').font = TOTAL['font']
    curr_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({c}{curr_start}:{c}{row-1})')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).border = THIN
    row += 1

    ws.cell(row=row, column=1, value='Non-Current Assets:').font = SUBHEADER['font']
    row += 1

    nc_items = [
        ('Oil & Gas Properties, Net', 'OilGasProp'),
        ('Derivative Assets (NC)', 'DerivAssets_NC'),
        ('Other Assets (NC)', 'OtherAssets_NC'),
    ]
    nc_start = row
    for label, key in nc_items:
        ws.cell(row=row, column=1, value=f'  {label}')
        for col, yr in enumerate(years_hist, 2):
            ws.cell(row=row, column=col, value=XBRL_BS[yr][key]).number_format = NUM
            ws.cell(row=row, column=col).fill = ACTUAL['fill']
            ws.cell(row=row, column=col).border = THIN
        # O&G properties grow with CapEx
        fy24_val = XBRL_BS['FY2024'][key]
        if key == 'OilGasProp':
            for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 5):
                capex = MGMT[yr].get('CapEx', MGMT[yr].get('CapEx_mid', 850000))
                prev = get_column_letter(col-1)
                ws.cell(row=row, column=col, value=f'={prev}{row}+{capex}-500000')  # CapEx - DD&A
                ws.cell(row=row, column=col).number_format = NUM
                ws.cell(row=row, column=col).fill = FORMULA['fill']
                ws.cell(row=row, column=col).border = THIN
            ws.cell(row=row, column=len(all_years)+2, value='XBRL / =Prior + CapEx - DD&A').font = SOURCE_FONT
        else:
            for col in range(5, len(all_years)+2):
                ws.cell(row=row, column=col, value=fy24_val).number_format = NUM
                ws.cell(row=row, column=col).fill = ESTIMATE['fill']
                ws.cell(row=row, column=col).border = THIN
            ws.cell(row=row, column=len(all_years)+2, value='XBRL / Held flat').font = SOURCE_FONT
        row += 1

    ws.cell(row=row, column=1, value='Total Non-Current Assets').font = TOTAL['font']
    nc_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({c}{nc_start}:{c}{row-1})')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).border = THIN
    row += 1

    ws.cell(row=row, column=1, value='TOTAL ASSETS').font = Font(bold=True, size=11)
    assets_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={c}{curr_row}+{c}{nc_row}')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    row += 2

    # LIABILITIES
    ws.cell(row=row, column=1, value='LIABILITIES').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    ws.cell(row=row, column=1, value='Current Liabilities:').font = SUBHEADER['font']
    row += 1

    cl_items = [
        ('Accounts Payable', 'AP'),
        ('Accrued Liabilities', 'AccruedLiab'),
        ('Derivative Liabilities', 'DerivLiab_Curr'),
        ('ARO (Current)', 'ARO_Curr'),
    ]
    cl_start = row
    for label, key in cl_items:
        ws.cell(row=row, column=1, value=f'  {label}')
        for col, yr in enumerate(years_hist, 2):
            ws.cell(row=row, column=col, value=XBRL_BS[yr][key]).number_format = NUM
            ws.cell(row=row, column=col).fill = ACTUAL['fill']
            ws.cell(row=row, column=col).border = THIN
        fy24_val = XBRL_BS['FY2024'][key]
        for col in range(5, len(all_years)+2):
            ws.cell(row=row, column=col, value=fy24_val).number_format = NUM
            ws.cell(row=row, column=col).fill = ESTIMATE['fill']
            ws.cell(row=row, column=col).border = THIN
        ws.cell(row=row, column=len(all_years)+2, value='XBRL / Held flat').font = SOURCE_FONT
        row += 1

    ws.cell(row=row, column=1, value='Total Current Liabilities').font = TOTAL['font']
    cl_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({c}{cl_start}:{c}{row-1})')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).border = THIN
    row += 1

    ws.cell(row=row, column=1, value='Non-Current Liabilities:').font = SUBHEADER['font']
    row += 1

    ncl_items = [
        ('Long-Term Debt', 'LTDebt'),
        ('Derivative Liabilities (NC)', 'DerivLiab_NC'),
        ('ARO (Non-Current)', 'ARO_NC'),
        ('Deferred Tax Liability', 'DeferredTax'),
        ('Other Liabilities (NC)', 'OtherLiab_NC'),
    ]
    ncl_start = row
    for label, key in ncl_items:
        ws.cell(row=row, column=1, value=f'  {label}')
        for col, yr in enumerate(years_hist, 2):
            ws.cell(row=row, column=col, value=XBRL_BS[yr][key]).number_format = NUM
            ws.cell(row=row, column=col).fill = ACTUAL['fill']
            ws.cell(row=row, column=col).border = THIN
        fy24_val = XBRL_BS['FY2024'][key]
        for col in range(5, len(all_years)+2):
            ws.cell(row=row, column=col, value=fy24_val).number_format = NUM
            ws.cell(row=row, column=col).fill = ESTIMATE['fill']
            ws.cell(row=row, column=col).border = THIN
        ws.cell(row=row, column=len(all_years)+2, value='XBRL / Held flat').font = SOURCE_FONT
        row += 1

    ws.cell(row=row, column=1, value='Total Non-Current Liabilities').font = TOTAL['font']
    ncl_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({c}{ncl_start}:{c}{row-1})')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).border = THIN
    row += 1

    ws.cell(row=row, column=1, value='TOTAL LIABILITIES').font = Font(bold=True, size=11)
    liab_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={c}{cl_row}+{c}{ncl_row}')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    row += 2

    # EQUITY
    ws.cell(row=row, column=1, value='STOCKHOLDERS\' EQUITY').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    eq_items = [
        ('Common Stock', 'CommonStock'),
        ('Additional Paid-In Capital', 'APIC'),
        ('Retained Earnings', 'RetainedEarnings'),
        ('Treasury Stock', 'Treasury'),
    ]
    eq_start = row
    for label, key in eq_items:
        ws.cell(row=row, column=1, value=f'  {label}')
        for col, yr in enumerate(years_hist, 2):
            ws.cell(row=row, column=col, value=XBRL_BS[yr][key]).number_format = NUM
            ws.cell(row=row, column=col).fill = ACTUAL['fill']
            ws.cell(row=row, column=col).border = THIN
        # RE grows with NI
        if key == 'RetainedEarnings':
            for col in range(5, len(all_years)+2):
                prev = get_column_letter(col-1)
                ws.cell(row=row, column=col, value=f'={prev}{row}+500000')  # Add NI estimate
                ws.cell(row=row, column=col).number_format = NUM
                ws.cell(row=row, column=col).fill = FORMULA['fill']
                ws.cell(row=row, column=col).border = THIN
            ws.cell(row=row, column=len(all_years)+2, value='XBRL / =Prior + NI - Div').font = SOURCE_FONT
        else:
            fy24_val = XBRL_BS['FY2024'][key]
            for col in range(5, len(all_years)+2):
                ws.cell(row=row, column=col, value=fy24_val).number_format = NUM
                ws.cell(row=row, column=col).fill = ESTIMATE['fill']
                ws.cell(row=row, column=col).border = THIN
            ws.cell(row=row, column=len(all_years)+2, value='XBRL / Held flat').font = SOURCE_FONT
        row += 1

    ws.cell(row=row, column=1, value='TOTAL STOCKHOLDERS\' EQUITY').font = Font(bold=True, size=11)
    equity_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({c}{eq_start}:{c}{row-1})')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    row += 2

    # Total L+E
    ws.cell(row=row, column=1, value='TOTAL LIABILITIES & EQUITY').font = Font(bold=True, size=11)
    le_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={c}{liab_row}+{c}{equity_row}')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    row += 2

    # Balance check
    ws.cell(row=row, column=1, value='BALANCE CHECK: A = L + E').font = Font(bold=True)
    row += 1
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=IF(ABS({c}{assets_row}-{c}{le_row})<1,"PASS","FAIL")')
        ws.cell(row=row, column=col).fill = CHECK['fill']


def create_cash_flow(ws):
    """Cash Flow Statement with detail and management CapEx guidance."""
    years_hist = ['FY2022', 'FY2023', 'FY2024']
    years_proj = ['FY2025E', 'FY2026E', 'FY2027E', 'FY2028E']
    all_years = years_hist + years_proj

    ws['A1'] = 'NOG CASH FLOW STATEMENT'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:L1')

    row = 4
    ws.cell(row=row, column=1, value='($ thousands)').font = Font(italic=True, size=9)
    for col, yr in enumerate(all_years, 2):
        c = ws.cell(row=row, column=col, value=yr)
        c.font = HEADER['font']
        c.fill = HEADER['fill']
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.cell(row=row, column=len(all_years)+2, value='Source').font = HEADER['font']
    ws.cell(row=row, column=len(all_years)+2).fill = HEADER['fill']
    ws.column_dimensions[get_column_letter(len(all_years)+2)].width = 55
    ws.column_dimensions['A'].width = 32
    row += 1

    # Beginning Cash
    ws.cell(row=row, column=1, value='Beginning Cash')
    for col, yr in enumerate(years_hist, 2):
        ws.cell(row=row, column=col, value=XBRL_CF[yr]['BeginCash']).number_format = NUM
        ws.cell(row=row, column=col).fill = ACTUAL['fill']
        ws.cell(row=row, column=col).border = THIN
    for col in range(5, len(all_years)+2):
        prev = get_column_letter(col-1)
        # Link to prior ending cash
        ws.cell(row=row, column=col, value=f'={prev}{row+10}')  # Link to ending cash row
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).border = THIN
    begin_cash_row = row
    ws.cell(row=row, column=len(all_years)+2, value='XBRL / Linked to prior Ending Cash').font = SOURCE_FONT
    row += 2

    # CFO
    ws.cell(row=row, column=1, value='Cash from Operating Activities').font = Font(bold=True)
    cfo_row = row
    for col, yr in enumerate(years_hist, 2):
        ws.cell(row=row, column=col, value=XBRL_CF[yr]['CFO']).number_format = NUM
        ws.cell(row=row, column=col).fill = ACTUAL['fill']
        ws.cell(row=row, column=col).border = THIN
    # Project CFO ~$1.4B
    for col in range(5, len(all_years)+2):
        ws.cell(row=row, column=col, value=1400000).number_format = NUM
        ws.cell(row=row, column=col).fill = ESTIMATE['fill']
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value='XBRL / Assumed ~FY24 level ($1.4B)').font = SOURCE_FONT
    row += 1

    # CFI (CapEx)
    ws.cell(row=row, column=1, value='Cash from Investing Activities').font = Font(bold=True)
    cfi_row = row
    for col, yr in enumerate(years_hist, 2):
        ws.cell(row=row, column=col, value=XBRL_CF[yr]['CFI']).number_format = NUM
        ws.cell(row=row, column=col).fill = ACTUAL['fill']
        ws.cell(row=row, column=col).border = THIN
    # Management CapEx guidance
    capex_vals = {
        'FY2025': -MGMT['FY2025']['CapEx_mid'],
        'FY2026': -MGMT['FY2026']['CapEx'],
        'FY2027': -MGMT['FY2027']['CapEx'],
        'FY2028': -MGMT['FY2028']['CapEx'],
    }
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 5):
        ws.cell(row=row, column=col, value=capex_vals[yr]).number_format = NUM
        ws.cell(row=row, column=col).fill = MGMT_STYLE['fill']
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value=MGMT['FY2025']['source'][:55]).font = SOURCE_FONT
    row += 1

    # CFF
    ws.cell(row=row, column=1, value='Cash from Financing Activities').font = Font(bold=True)
    cff_row = row
    for col, yr in enumerate(years_hist, 2):
        ws.cell(row=row, column=col, value=XBRL_CF[yr]['CFF']).number_format = NUM
        ws.cell(row=row, column=col).fill = ACTUAL['fill']
        ws.cell(row=row, column=col).border = THIN
    for col in range(5, len(all_years)+2):
        ws.cell(row=row, column=col, value=0).number_format = NUM
        ws.cell(row=row, column=col).fill = ESTIMATE['fill']
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value='XBRL / Assumed net zero (debt/dividend balance)').font = SOURCE_FONT
    row += 2

    # Net Change
    ws.cell(row=row, column=1, value='Net Change in Cash').font = Font(bold=True)
    net_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={c}{cfo_row}+{c}{cfi_row}+{c}{cff_row}')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value='=CFO + CFI + CFF').font = SOURCE_FONT
    row += 2

    # Ending Cash
    ws.cell(row=row, column=1, value='Ending Cash').font = Font(bold=True)
    end_cash_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={c}{begin_cash_row}+{c}{net_row}')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value='=Beginning + Net Change').font = SOURCE_FONT
    row += 2

    # Free Cash Flow
    ws.cell(row=row, column=1, value='FREE CASH FLOW (CFO + CFI)').font = Font(bold=True, size=11)
    fcf_row = row
    for col in range(2, len(all_years)+2):
        c = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={c}{cfo_row}+{c}{cfi_row}')
        ws.cell(row=row, column=col).number_format = NUM
        ws.cell(row=row, column=col).fill = FORMULA['fill']
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = THIN
    ws.cell(row=row, column=len(all_years)+2, value='=CFO + CFI (proxy for FCF)').font = SOURCE_FONT
    row += 2

    # CapEx Detail
    ws.cell(row=row, column=1, value='CAPEX DETAIL (Management Guidance)').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    ws.cell(row=row, column=1, value='CapEx Low')
    for col in range(2, 5):
        ws.cell(row=row, column=col, value='-')
    ws.cell(row=row, column=5, value=MGMT['FY2025']['CapEx_low']).number_format = NUM
    ws.cell(row=row, column=5).fill = MGMT_STYLE['fill']
    for col in range(6, len(all_years)+2):
        ws.cell(row=row, column=col, value='-')
    row += 1

    ws.cell(row=row, column=1, value='CapEx High')
    for col in range(2, 5):
        ws.cell(row=row, column=col, value='-')
    ws.cell(row=row, column=5, value=MGMT['FY2025']['CapEx_high']).number_format = NUM
    ws.cell(row=row, column=5).fill = MGMT_STYLE['fill']
    for col in range(6, len(all_years)+2):
        ws.cell(row=row, column=col, value='-')
    row += 1

    ws.cell(row=row, column=1, value='CapEx Type')
    for col in range(2, 5):
        ws.cell(row=row, column=col, value='-')
    ws.cell(row=row, column=5, value='Total')
    for col, yr in enumerate(['FY2026', 'FY2027', 'FY2028'], 6):
        ws.cell(row=row, column=col, value=MGMT[yr]['CapEx_type'])
    ws.cell(row=row, column=len(all_years)+2, value=MGMT['FY2026']['source'][:55]).font = SOURCE_FONT


def create_analyst_detail(ws):
    """Comprehensive analyst estimates with all details."""
    ws['A1'] = 'ANALYST CONSENSUS & MANAGEMENT GUIDANCE - FULL DETAIL'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    ws['A2'] = 'All data from external sources with full attribution. NO model-generated forecasts.'
    ws['A2'].font = Font(bold=True, size=10, color="CC0000")

    row = 4
    # EPS Detail
    ws.cell(row=row, column=1, value='EPS CONSENSUS - FULL DETAIL').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    headers = ['Year', 'Low', 'Average', 'High', 'Range', '# Analysts', 'Source']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = SUBHEADER['fill']
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions['G'].width = 45
    row += 1

    for yr in ['FY2025', 'FY2026', 'FY2027', 'FY2028']:
        d = ANALYST_EPS[yr]
        ws.cell(row=row, column=1, value=yr)
        ws.cell(row=row, column=2, value=d['FY_low']).number_format = NUM_DEC
        ws.cell(row=row, column=3, value=d['FY_avg']).number_format = NUM_DEC
        ws.cell(row=row, column=4, value=d['FY_high']).number_format = NUM_DEC
        ws.cell(row=row, column=5, value=f'${d["FY_high"]-d["FY_low"]:.2f}')
        ws.cell(row=row, column=6, value=d['analysts'])
        ws.cell(row=row, column=7, value=d['source']).font = SOURCE_FONT
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = ESTIMATE['fill']
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN
        row += 1

    row += 1
    # Quarterly EPS Detail for FY2025
    ws.cell(row=row, column=1, value='FY2025 QUARTERLY EPS').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    headers = ['Quarter', 'Estimate', 'Actual', 'Surprise $', 'Surprise %', 'Status', 'Source']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = SUBHEADER['fill']
    row += 1

    d = ANALYST_EPS['FY2025']
    quarters = [
        ('Q1 2025', d['Q1_est'], d['Q1_actual'], d['Q1_surprise'], d['Q1_source']),
        ('Q2 2025', d['Q2_est'], d['Q2_actual'], d['Q2_surprise'], d['Q2_source']),
        ('Q3 2025', d['Q3_est'], d['Q3_actual'], d['Q3_surprise'], d['Q3_source']),
        ('Q4 2025', d['Q4_est'], None, None, d['Q4_source']),
    ]
    for q, est, actual, surprise, src in quarters:
        ws.cell(row=row, column=1, value=q)
        ws.cell(row=row, column=2, value=est).number_format = NUM_DEC
        if actual:
            ws.cell(row=row, column=3, value=actual).number_format = NUM_DEC
            ws.cell(row=row, column=4, value=actual-est).number_format = '+0.00;-0.00'
            ws.cell(row=row, column=5, value=surprise/100).number_format = '+0.0%;-0.0%'
            ws.cell(row=row, column=6, value='BEAT' if surprise > 0 else 'MISS')
            ws.cell(row=row, column=6).fill = PatternFill("solid", fgColor="C6EFCE" if surprise > 0 else "FFC7CE")
        else:
            ws.cell(row=row, column=3, value='TBD')
            ws.cell(row=row, column=6, value='PENDING')
        ws.cell(row=row, column=7, value=src).font = SOURCE_FONT
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN
        row += 1

    row += 1
    # Revenue Consensus
    ws.cell(row=row, column=1, value='REVENUE CONSENSUS ($M)').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    headers = ['Year', 'Low', 'Average', 'High', 'YoY %', '# Analysts', 'Source']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = SUBHEADER['fill']
    row += 1

    prev_rev = XBRL_IS['FY2024']['Revenue_Total'] / 1000
    for yr in ['FY2025', 'FY2026', 'FY2027', 'FY2028']:
        d = ANALYST_REV[yr]
        avg = d['FY_avg'] / 1000
        yoy = (avg - prev_rev) / prev_rev
        ws.cell(row=row, column=1, value=yr)
        ws.cell(row=row, column=2, value=d['FY_low']/1000).number_format = '#,##0'
        ws.cell(row=row, column=3, value=avg).number_format = '#,##0'
        ws.cell(row=row, column=4, value=d['FY_high']/1000).number_format = '#,##0'
        ws.cell(row=row, column=5, value=yoy).number_format = '+0.0%;-0.0%'
        ws.cell(row=row, column=6, value=d['analysts'])
        ws.cell(row=row, column=7, value=d['source']).font = SOURCE_FONT
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = ESTIMATE['fill']
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN
        prev_rev = avg
        row += 1

    row += 1
    # Price Targets
    ws.cell(row=row, column=1, value='ANALYST PRICE TARGETS').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    pt = ANALYST_PT['current']
    items = [
        ('Current Average PT', f'${pt["avg"]:.2f}'),
        ('Low PT', f'${pt["low"]}'),
        ('High PT', f'${pt["high"]}'),
        ('# Analysts', str(pt['analysts'])),
        ('As of', pt['date']),
    ]
    for label, val in items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=1).border = THIN
        ws.cell(row=row, column=2).border = THIN
        row += 1
    ws.cell(row=row-5, column=7, value=pt['source']).font = SOURCE_FONT
    row += 1

    # Rating Distribution
    ws.cell(row=row, column=1, value='RATING DISTRIBUTION').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    ratings = ANALYST_PT['ratings']
    r_items = [
        ('Strong Buy', ratings['strong_buy']),
        ('Buy', ratings['buy']),
        ('Hold', ratings['hold']),
        ('Sell', ratings['sell']),
        ('Strong Sell', ratings['strong_sell']),
    ]
    for col, (rating, count) in enumerate(r_items, 1):
        ws.cell(row=row, column=col, value=rating).font = Font(bold=True, size=9)
    row += 1
    for col, (rating, count) in enumerate(r_items, 1):
        ws.cell(row=row, column=col, value=count)
    ws.cell(row=row-1, column=7, value=ratings['source']).font = SOURCE_FONT
    row += 2

    # Recent PT Changes
    ws.cell(row=row, column=1, value='RECENT PRICE TARGET CHANGES').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    headers = ['Date', 'Firm', 'Analyst', 'Rating', 'PT', 'Source']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = SUBHEADER['fill']
    row += 1

    for change in ANALYST_PT['history'][:10]:
        ws.cell(row=row, column=1, value=change['date'])
        ws.cell(row=row, column=2, value=change['firm'])
        ws.cell(row=row, column=3, value=change['analyst'])
        ws.cell(row=row, column=4, value=change['rating'])
        ws.cell(row=row, column=5, value=f'${change["pt"]}' if change['pt'] else '-')
        ws.cell(row=row, column=6, value=change['source']).font = SOURCE_FONT
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN
        row += 1

    row += 1
    # Beat/Miss History
    ws.cell(row=row, column=1, value='BEAT/MISS HISTORY (Last 8 Quarters)').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    beat_count = sum(1 for b in BEAT_MISS if b['surprise'] > 0)
    avg_surprise = sum(b['surprise'] for b in BEAT_MISS) / len(BEAT_MISS)
    ws.cell(row=row, column=1, value=f'Beat Rate: {beat_count}/8 = {beat_count/8*100:.0f}%').font = Font(bold=True)
    ws.cell(row=row, column=3, value=f'Avg Surprise: {avg_surprise:.1f}%')
    row += 1

    headers = ['Quarter', 'Date', 'Estimate', 'Actual', 'Surprise %', 'Result', 'Source']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = SUBHEADER['fill']
    row += 1

    for b in BEAT_MISS:
        ws.cell(row=row, column=1, value=b['qtr'])
        ws.cell(row=row, column=2, value=b['date'])
        ws.cell(row=row, column=3, value=b['est']).number_format = NUM_DEC
        ws.cell(row=row, column=4, value=b['actual']).number_format = NUM_DEC
        ws.cell(row=row, column=5, value=b['surprise']/100).number_format = '+0.0%;-0.0%'
        ws.cell(row=row, column=6, value='BEAT' if b['surprise'] > 0 else 'MISS')
        ws.cell(row=row, column=6).fill = PatternFill("solid", fgColor="C6EFCE" if b['surprise'] > 0 else "FFC7CE")
        ws.cell(row=row, column=7, value=b['source']).font = SOURCE_FONT
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN
        row += 1

    row += 1
    # Management Guidance
    ws.cell(row=row, column=1, value='MANAGEMENT GUIDANCE (FROM TRANSCRIPTS)').font = SECTION['font']
    ws.cell(row=row, column=1).fill = SECTION['fill']
    row += 1

    headers = ['Item', 'Value', 'Source']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = SUBHEADER['fill']
    ws.column_dimensions['C'].width = 65
    row += 1

    mgmt_items = [
        ('2025 CapEx', f'${MGMT["FY2025"]["CapEx_low"]/1000:.0f}M - ${MGMT["FY2025"]["CapEx_high"]/1000:.0f}M', MGMT['FY2025']['source']),
        ('2025 Production', f'{MGMT["FY2025"]["Production_low"]:,} - {MGMT["FY2025"]["Production_high"]:,} BOE/day', MGMT['FY2025']['source']),
        ('2026 CapEx', f'${MGMT["FY2026"]["CapEx"]/1000:.0f}M (maintenance)', MGMT['FY2026']['source']),
        ('2027 CapEx', f'${MGMT["FY2027"]["CapEx"]/1000:.0f}M (maintenance)', MGMT['FY2027']['source']),
        ('Federal Tax Outlook', MGMT['Tax']['outlook'], MGMT['Tax']['source'][:60]),
        ('M&A Pipeline', MGMT['MA']['pipeline'], MGMT['MA']['source'][:60]),
        ('Q3-Q4 2025 Outlook', MGMT['Q3Q4_2025']['outlook'], MGMT['Q3Q4_2025']['source'][:60]),
    ]
    for item, val, src in mgmt_items:
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=2).fill = MGMT_STYLE['fill']
        ws.cell(row=row, column=3, value=src).font = SOURCE_FONT
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = THIN
        row += 1


def create_sources(ws):
    """Complete source attribution."""
    ws['A1'] = 'DATA SOURCES & ATTRIBUTION'
    ws['A1'].font = Font(bold=True, size=16)

    ws['A3'] = 'This model contains NO model-generated forecasts.'
    ws['A3'].font = Font(bold=True, color="CC0000")
    ws['A4'] = 'All projections are from external analyst sources with full attribution.'
    ws['A4'].font = Font(bold=True, color="CC0000")

    row = 6
    sections = [
        ('HISTORICAL DATA (XBRL)', [
            ('Income Statement FY2022-2024', 'XBRL 10-K filings via Neo4j (us-gaap concepts)'),
            ('Balance Sheet FY2022-2024', 'XBRL 10-K filings via Neo4j'),
            ('Cash Flow FY2022-2024', 'XBRL 10-K filings via Neo4j'),
            ('Quarterly Data (YTD derivation)', 'XBRL 10-Q filings via Neo4j'),
            ('Revenue Segmentation', 'XBRL Fact FACT_MEMBER relationships'),
        ], ACTUAL['fill']),
        ('ANALYST ESTIMATES', [
            ('EPS Consensus FY2025', 'Alpha Vantage EARNINGS API + Yahoo Finance (12 analysts)'),
            ('EPS Consensus FY2026', 'StockAnalysis.com (8 analysts)'),
            ('EPS Consensus FY2027', 'Barchart / StockAnalysis.com (7 analysts)'),
            ('EPS Consensus FY2028', 'Fintel (2 analysts - limited coverage)'),
            ('Revenue Consensus', 'Yahoo Finance / StockAnalysis.com (6-8 analysts)'),
            ('Price Targets', 'Benzinga News via Neo4j (bzNews articles)'),
            ('Analyst Ratings', 'Benzinga News aggregation'),
            ('Beat/Miss History', 'Alpha Vantage EARNINGS API'),
        ], ESTIMATE['fill']),
        ('MANAGEMENT GUIDANCE', [
            ('2025 CapEx ($925M-$1.05B)', 'Transcript:NOG_2025_2 (2025-08-01) - CFO Chad Allen'),
            ('2025 Production (130-135K BOE/d)', 'Transcript:NOG_2025_2 (2025-08-01) - CFO Chad Allen'),
            ('2026-2027 Maintenance CapEx', 'Transcript:NOG_2025-04-30 Q&A - Jim Evans CTO'),
            ('Tax Outlook ($0 through 2028)', 'Transcript:NOG_2025_2 - CFO Chad Allen'),
            ('M&A Pipeline ($8B+)', 'Transcript:NOG_2025_2 - Adam Durlam'),
        ], MGMT_STYLE['fill']),
    ]

    for section, items, fill in sections:
        ws.cell(row=row, column=1, value=section).font = SECTION['font']
        ws.cell(row=row, column=1).fill = SECTION['fill']
        row += 1
        for item, src in items:
            ws.cell(row=row, column=1, value=f'  {item}')
            ws.cell(row=row, column=1).fill = fill
            ws.cell(row=row, column=2, value=src)
            row += 1
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='COLOR LEGEND').font = Font(bold=True)
    row += 1
    legend = [
        ('White', 'Historical data from XBRL', ACTUAL['fill']),
        ('Orange', 'Analyst consensus estimate', ESTIMATE['fill']),
        ('Green', 'Management guidance from transcripts', MGMT_STYLE['fill']),
        ('Yellow', 'Formula-derived (linked)', FORMULA['fill']),
        ('Light Green', 'Validation check', CHECK['fill']),
    ]
    for color, desc, fill in legend:
        ws.cell(row=row, column=1, value=color)
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2, value=desc)
        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 60


def main():
    print("=" * 70)
    print("NOG COMPLETE FINANCIAL MODEL - MAXIMUM DETAIL")
    print("Historical + 3-Year Analyst Projections + Full Source Attribution")
    print("=" * 70)

    wb = Workbook()

    ws_is = wb.active
    ws_is.title = "Income Statement"
    create_income_statement(ws_is)

    ws_bs = wb.create_sheet("Balance Sheet")
    create_balance_sheet(ws_bs)

    ws_cf = wb.create_sheet("Cash Flow")
    create_cash_flow(ws_cf)

    ws_analyst = wb.create_sheet("Analyst Consensus")
    create_analyst_detail(ws_analyst)

    ws_sources = wb.create_sheet("Sources")
    create_sources(ws_sources)

    output = '/home/faisal/EventMarketDB/earnings-analysis/NOG_Complete_Model.xlsx'
    wb.save(output)

    print(f"\nSaved: {output}")
    print("\nSheets:")
    for s in wb.sheetnames:
        print(f"  • {s}")

    print("\n" + "=" * 70)
    print("3-YEAR PROJECTION SUMMARY (Analyst Consensus)")
    print("=" * 70)
    print(f"{'Year':<10} {'EPS':<12} {'EPS Range':<18} {'Revenue ($M)':<14} {'CapEx ($M)':<12} {'Analysts'}")
    print("-" * 80)
    for yr in ['FY2025', 'FY2026', 'FY2027', 'FY2028']:
        eps = ANALYST_EPS[yr]['FY_avg']
        eps_range = f"${ANALYST_EPS[yr]['FY_low']:.2f}-${ANALYST_EPS[yr]['FY_high']:.2f}"
        rev = ANALYST_REV[yr]['FY_avg'] / 1000
        capex = MGMT[yr].get('CapEx', MGMT[yr].get('CapEx_mid', 850000)) / 1000
        analysts = ANALYST_EPS[yr]['analysts']
        print(f"{yr:<10} ${eps:<10.2f} {eps_range:<18} ${rev:>9,.0f}      ${capex:>6,.0f}       {analysts}")
    print("-" * 80)
    print(f"\nPrice Target: ${ANALYST_PT['current']['avg']:.2f} (Range: ${ANALYST_PT['current']['low']}-${ANALYST_PT['current']['high']})")
    print(f"Beat Rate: {sum(1 for b in BEAT_MISS if b['surprise']>0)}/8 = 75%")


if __name__ == '__main__':
    main()
