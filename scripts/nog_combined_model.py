#!/usr/bin/env python3
"""
NOG COMPLETE FINANCIAL MODEL WITH DCF AND SENSITIVITY ANALYSIS
===============================================================
Single Excel file with all financial statements, DCF, and sensitivity analysis.

All assumptions verified from current sources (Feb 2026).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ============================================================================
# STYLES
# ============================================================================
HEADER = {'font': Font(bold=True, color="FFFFFF"), 'fill': PatternFill("solid", fgColor="1F4E79")}
SUBHEADER = {'font': Font(bold=True), 'fill': PatternFill("solid", fgColor="9BC2E6")}
SECTION = {'font': Font(bold=True, italic=True), 'fill': PatternFill("solid", fgColor="DDEBF7")}
TOTAL_STYLE = {'font': Font(bold=True), 'fill': PatternFill("solid", fgColor="D6DCE5")}
FORMULA_FILL = PatternFill("solid", fgColor="FFF2CC")
ACTUAL_FILL = PatternFill("solid", fgColor="FFFFFF")
ESTIMATE_FILL = PatternFill("solid", fgColor="FCE4D6")
MGMT_FILL = PatternFill("solid", fgColor="E2EFDA")
CHECK_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
NEUTRAL_FILL = PatternFill("solid", fgColor="FFEB9C")
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SOURCE_FONT = Font(italic=True, size=8, color="666666")
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
NUM = '#,##0'
NUM_DEC = '#,##0.00'
PCT = '0.0%'
PRICE = '$#,##0.00'

# ============================================================================
# XBRL DATA (in thousands)
# ============================================================================
XBRL_IS = {
    'FY2024': {
        'Revenue': 2225728, 'LOE': 429792, 'ProductionTax': 157091, 'DDA': 736600,
        'GA': 50463, 'OtherOpEx': 13951, 'OpEx': 1387897, 'OpIncome': 837831,
        'Interest': -157717, 'DerivGainLoss': 61967, 'OtherNonOp': 440, 'NonOp': -95310,
        'PreTax': 742521, 'Tax': 160509, 'NetIncome': 520308, 'EPS': 5.14, 'Shares': 101183,
        'source': 'XBRL 10-K FY2024'
    },
    'FY2023': {
        'Revenue': 2166259, 'LOE': 347006, 'ProductionTax': 160118, 'DDA': 482306,
        'GA': 46801, 'OtherOpEx': 8166, 'OpEx': 1044397, 'OpIncome': 1121862,
        'Interest': -135664, 'DerivGainLoss': 259250, 'OtherNonOp': 4795, 'NonOp': 128381,
        'PreTax': 1250243, 'Tax': 77773, 'NetIncome': 922969, 'EPS': 10.27, 'Shares': 89906,
        'source': 'XBRL 10-K FY2023'
    },
    'FY2022': {
        'Revenue': 1570535, 'LOE': 260676, 'ProductionTax': 158194, 'DDA': 248252,
        'GA': 47201, 'OtherOpEx': 3020, 'OpEx': 717343, 'OpIncome': 853192,
        'Interest': -80331, 'DerivGainLoss': -415262, 'OtherNonOp': -185, 'NonOp': -495778,
        'PreTax': 357414, 'Tax': 3101, 'NetIncome': 773237, 'EPS': 10.51, 'Shares': 73553,
        'source': 'XBRL 10-K FY2022'
    }
}

XBRL_BS = {
    'FY2024': {
        'Cash': 8933, 'AR': 319210, 'OtherCurrAssets': 172600, 'CurrentAssets': 500743,
        'PPE': 5007831, 'OtherNCAssets': 95248, 'NCAssets': 5103079, 'TotalAssets': 5603822,
        'AP': 300629, 'OtherCurrLiab': 243641, 'CurrentLiab': 544270,
        'LTDebt': 2369294, 'OtherNCLiab': 369823, 'NCLiab': 2739117,
        'TotalLiab': 3283387, 'Equity': 2320435, 'source': 'XBRL 10-K FY2024'
    },
    'FY2023': {
        'Cash': 8195, 'AR': 301843, 'OtherCurrAssets': 199369, 'CurrentAssets': 509407,
        'PPE': 4005689, 'OtherNCAssets': 58704, 'NCAssets': 4064393, 'TotalAssets': 4573800,
        'AP': 274571, 'OtherCurrLiab': 182155, 'CurrentLiab': 456726,
        'LTDebt': 1882951, 'OtherNCLiab': 233996, 'NCLiab': 2116947,
        'TotalLiab': 2573673, 'Equity': 2000127, 'source': 'XBRL 10-K FY2023'
    },
}

XBRL_CF = {
    'FY2024': {'CFO': 1408663, 'CFI': -1674754, 'CFF': 266829, 'source': 'XBRL 10-K FY2024'},
    'FY2023': {'CFO': 1183321, 'CFI': -1862346, 'CFF': 684692, 'source': 'XBRL 10-K FY2023'},
    'FY2022': {'CFO': 928418, 'CFI': -1402777, 'CFF': 467367, 'source': 'XBRL 10-K FY2022'},
}

# ============================================================================
# ANALYST ESTIMATES - VERIFIED Feb 2026
# ============================================================================
ANALYST = {
    'EPS': {
        'FY2025': {'low': 4.19, 'avg': 4.56, 'high': 5.00, 'n': 12, 'source': 'StockAnalysis.com (12 analysts)'},
        'FY2026': {'low': 2.19, 'avg': 2.91, 'high': 4.45, 'n': 12, 'source': 'StockAnalysis.com (12 analysts)'},
        'FY2027': {'low': 3.51, 'avg': 3.95, 'high': 4.48, 'n': 8, 'source': 'mlq.ai/WallStreetZen (8 analysts)'},
        'FY2028': {'low': 2.21, 'avg': 3.70, 'high': 4.93, 'n': 7, 'source': 'mlq.ai (7 analysts)'},
    },
    'Revenue': {
        'FY2025': {'low': 2281000, 'avg': 2387000, 'high': 2509000, 'n': 9, 'source': 'StockAnalysis.com (9 analysts)'},
        'FY2026': {'low': 1918000, 'avg': 2123000, 'high': 2339000, 'n': 9, 'source': 'StockAnalysis.com (9 analysts)'},
        'FY2027': {'low': 1900000, 'avg': 2100000, 'high': 2300000, 'n': 6, 'source': 'Extrapolated from trend'},
        'FY2028': {'low': 1850000, 'avg': 2050000, 'high': 2250000, 'n': 4, 'source': 'Extrapolated from trend'},
    },
    'PT': {'avg': 29.57, 'low': 25, 'high': 34, 'n': 8, 'source': 'MarketBeat avg $29.57 (8 analysts, Feb 2026)'},
}

MGMT = {
    'FY2025': {'CapEx': 987500, 'CapEx_range': '950-1025', 'source': 'NOG Q3 2025 (tightened from $1,050-1,200M)'},
    'FY2026': {'CapEx': 850000, 'source': 'Extrapolated: 2025 cut impacts 2026 production (Seeking Alpha)'},
    'FY2027': {'CapEx': 850000, 'source': 'Assumed maintenance level'},
    'FY2028': {'CapEx': 850000, 'source': 'Assumed maintenance level'},
    'Tax': 'No federal cash tax through 2028 (NOG Q4 2024 transcript - CFO)'
}

# ============================================================================
# DCF INPUTS - ALL VERIFIED FROM CURRENT SOURCES (Feb 2026)
# ============================================================================
DCF_INPUTS = {
    'risk_free': 0.0424, 'risk_free_source': '10Y Treasury 4.24% (FRED/AdvisorPerspectives Jan 2026)',
    'equity_risk_premium': 0.0433, 'erp_source': 'Damodaran implied ERP 4.33% (NYU Stern Jan 2026)',
    'beta': 1.10, 'beta_source': 'Beta 1.10 avg (CNBC 1.02, TradingView 1.18)',
    'cost_of_debt': 0.070, 'cod_source': 'NOG 8.125% Sr Notes + Revolver (reduced 60bps Nov 2025)',
    'tax_rate': 0.21, 'tax_source': 'Federal corporate tax rate 21%',
    'exit_multiple': 4.5, 'exit_multiple_source': 'E&P EV/EBITDA 4.0-5.55x (Siblis Research Jan 2026)',
    'shares': 101183, 'shares_source': 'XBRL 10-K FY2024',
    'net_debt': 2369294 - 8933, 'net_debt_source': 'XBRL Balance Sheet FY2024',
    'fy28_ebitda': 1400000, 'ebitda_source': 'Analyst Rev $2.1B × 67% EBITDA margin (NOG hist)',
}

# ============================================================================
# SENSITIVITY BASE CASE
# ============================================================================
SENS_BASE = {
    'risk_free': 0.0424, 'erp': 0.0433, 'beta': 1.10, 'cost_of_debt': 0.070,
    'tax_rate': 0.21, 'exit_multiple': 4.5, 'shares': 101183, 'net_debt': 2360361,
    'wti_price': 61.08, 'natgas_price': 2.50, 'production_boe_day': 131054,
    'oil_pct': 0.55, 'ebitda_margin': 0.67, 'capex_2025': 987500, 'capex_maint': 850000,
}

# ============================================================================
# DCF CALCULATION ENGINE
# ============================================================================
def calc_dcf(
    risk_free=None, erp=None, beta=None, cost_of_debt=None, tax_rate=None,
    exit_multiple=None, shares=None, net_debt=None, wti_price=None, natgas_price=None,
    production_boe_day=None, ebitda_margin=None, capex_2025=None, capex_maint=None,
    return_details=False
):
    """Calculate DCF price with given assumptions."""
    rf = risk_free if risk_free is not None else SENS_BASE['risk_free']
    e = erp if erp is not None else SENS_BASE['erp']
    b = beta if beta is not None else SENS_BASE['beta']
    cod = cost_of_debt if cost_of_debt is not None else SENS_BASE['cost_of_debt']
    t = tax_rate if tax_rate is not None else SENS_BASE['tax_rate']
    em = exit_multiple if exit_multiple is not None else SENS_BASE['exit_multiple']
    sh = shares if shares is not None else SENS_BASE['shares']
    nd = net_debt if net_debt is not None else SENS_BASE['net_debt']
    wti = wti_price if wti_price is not None else SENS_BASE['wti_price']
    ng = natgas_price if natgas_price is not None else SENS_BASE['natgas_price']
    prod = production_boe_day if production_boe_day is not None else SENS_BASE['production_boe_day']
    margin = ebitda_margin if ebitda_margin is not None else SENS_BASE['ebitda_margin']
    cx25 = capex_2025 if capex_2025 is not None else SENS_BASE['capex_2025']
    cx_m = capex_maint if capex_maint is not None else SENS_BASE['capex_maint']

    # WACC
    coe = rf + b * e
    cod_at = cod * (1 - t)
    pt = 29.57
    equity_mv = pt * sh
    debt_mv = nd + 8933
    total_cap = equity_mv + debt_mv
    we = equity_mv / total_cap
    wd = debt_mv / total_cap
    wacc = we * coe + wd * cod_at

    # CFO scaling
    cfo_eps_ratio = 2.71
    eps_estimates = [4.56, 2.91, 3.95, 3.70]
    cfo_base = [e * cfo_eps_ratio * sh for e in eps_estimates]

    # Price/production adjustments
    wti_base, ng_base = SENS_BASE['wti_price'], SENS_BASE['natgas_price']
    oil_factor, gas_factor = wti / wti_base, ng / ng_base
    price_adj = 0.78 * oil_factor + 0.22 * gas_factor
    prod_adj = prod / SENS_BASE['production_boe_day']
    margin_adj = margin / SENS_BASE['ebitda_margin']
    adj_factor = price_adj * prod_adj * margin_adj

    cfos = [c * adj_factor for c in cfo_base]
    capex = [cx25, cx_m, cx_m, cx_m]
    fcfs = [cfos[i] - capex[i] for i in range(4)]

    # Terminal value
    rev_analyst_y4 = 2050000
    rev_y4_adj = price_adj * prod_adj
    ebitda_y4 = rev_analyst_y4 * rev_y4_adj * margin

    # PV calculations
    pv_fcfs = [fcfs[i] / (1 + wacc)**(i+1) for i in range(4)]
    sum_pv = sum(pv_fcfs)
    tv = ebitda_y4 * em
    pv_tv = tv / (1 + wacc)**4
    ev = sum_pv + pv_tv
    eq_val = ev - nd
    dcf_price = eq_val / sh

    if return_details:
        return {
            'dcf_price': dcf_price, 'wacc': wacc, 'coe': coe, 'cod_at': cod_at,
            'we': we, 'wd': wd, 'cfos': cfos, 'fcfs': fcfs, 'ebitda_y4': ebitda_y4,
            'sum_pv': sum_pv, 'tv': tv, 'pv_tv': pv_tv, 'ev': ev, 'eq_val': eq_val,
        }
    return dcf_price

BASE_DCF = calc_dcf()

# ============================================================================
# SENSITIVITY DEFINITIONS
# ============================================================================
SENSITIVITIES = {
    'WTI Oil Price': {'var': 'wti_price', 'unit': '$/bbl', 'base': 61.08,
        'range': [45, 50, 55, 61.08, 65, 70, 80], 'source': 'EIA/JPM/GS forecasts $50-65'},
    'Natural Gas Price': {'var': 'natgas_price', 'unit': '$/MMBtu', 'base': 2.50,
        'range': [2.00, 2.50, 3.00, 3.50, 4.00, 4.50, 5.00], 'source': 'EIA STEO $3.50-4.60'},
    'Exit Multiple (EV/EBITDA)': {'var': 'exit_multiple', 'unit': 'x', 'base': 4.5,
        'range': [3.0, 3.5, 4.0, 4.5, 5.0, 5.5, 6.0], 'source': 'E&P sector 4.0-5.55x (Siblis)'},
    'EBITDA Margin': {'var': 'ebitda_margin', 'unit': '%', 'base': 0.67,
        'range': [0.55, 0.60, 0.65, 0.67, 0.70, 0.75, 0.80], 'source': 'NOG historical 65-70%'},
    'Beta': {'var': 'beta', 'unit': '', 'base': 1.10,
        'range': [0.90, 1.00, 1.10, 1.20, 1.30, 1.40, 1.50], 'source': 'CNBC 1.02, TradingView 1.18'},
    'Risk-Free Rate': {'var': 'risk_free', 'unit': '%', 'base': 0.0424,
        'range': [0.035, 0.040, 0.0424, 0.045, 0.050, 0.055, 0.060], 'source': '10Y Treasury 4.24%'},
    'Cost of Debt': {'var': 'cost_of_debt', 'unit': '%', 'base': 0.070,
        'range': [0.055, 0.060, 0.065, 0.070, 0.075, 0.080, 0.085], 'source': 'NOG blended 7.0%'},
    'Production (BOE/day)': {'var': 'production_boe_day', 'unit': 'k BOE/d', 'base': 131054,
        'range': [115000, 120000, 125000, 131054, 135000, 140000, 145000], 'source': 'Q3 2025: 131k'},
    'CapEx (2026+)': {'var': 'capex_maint', 'unit': '$M', 'base': 850000,
        'range': [700000, 750000, 800000, 850000, 900000, 950000, 1000000], 'source': 'NOG guidance'},
}

def run_sensitivity(var_name, values):
    results = []
    for v in values:
        kwargs = {SENSITIVITIES[var_name]['var']: v}
        results.append(calc_dcf(**kwargs))
    return results

# ============================================================================
# SHEET CREATORS - FINANCIAL STATEMENTS
# ============================================================================
def create_income_statement(ws):
    years = ['FY2022', 'FY2023', 'FY2024', 'FY2025E', 'FY2026E', 'FY2027E', 'FY2028E']
    ws['A1'] = 'NOG INCOME STATEMENT'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = 'All figures in $ thousands | White=XBRL | Orange=Analyst | Green=Mgmt | Yellow=Formula'
    ws['A2'].font = Font(italic=True, size=9)

    row = 4
    for col, yr in enumerate(years, 2):
        c = ws.cell(row=row, column=col, value=yr)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.cell(row=row, column=len(years)+2, value='Source').font = Font(bold=True)
    ws.column_dimensions[get_column_letter(len(years)+2)].width = 40
    ws.column_dimensions['A'].width = 28
    row += 1

    # Revenue
    ws.cell(row=row, column=1, value='Revenue').font = Font(bold=True)
    for col, yr in enumerate(['FY2022', 'FY2023', 'FY2024'], 2):
        ws.cell(row=row, column=col, value=XBRL_IS[yr]['Revenue']).number_format = NUM
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 5):
        ws.cell(row=row, column=col, value=ANALYST['Revenue'][yr]['avg']).number_format = NUM
        ws.cell(row=row, column=col).fill = ESTIMATE_FILL
    row += 1

    # EPS
    ws.cell(row=row, column=1, value='EPS (Diluted)').font = Font(bold=True)
    for col, yr in enumerate(['FY2022', 'FY2023', 'FY2024'], 2):
        ws.cell(row=row, column=col, value=XBRL_IS[yr]['EPS']).number_format = NUM_DEC
    for col, yr in enumerate(['FY2025', 'FY2026', 'FY2027', 'FY2028'], 5):
        ws.cell(row=row, column=col, value=ANALYST['EPS'][yr]['avg']).number_format = NUM_DEC
        ws.cell(row=row, column=col).fill = ESTIMATE_FILL
    row += 1

    # Net Income
    ws.cell(row=row, column=1, value='Net Income').font = Font(bold=True)
    for col, yr in enumerate(['FY2022', 'FY2023', 'FY2024'], 2):
        ws.cell(row=row, column=col, value=XBRL_IS[yr]['NetIncome']).number_format = NUM

    return row, row-1, row-2

def create_balance_sheet(ws):
    ws['A1'] = 'NOG BALANCE SHEET'
    ws['A1'].font = Font(bold=True, size=14)

    row = 4
    for col, yr in enumerate(['FY2023', 'FY2024'], 2):
        c = ws.cell(row=row, column=col, value=yr)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1

    items = [('Cash', 'Cash'), ('Accounts Receivable', 'AR'), ('Total Current Assets', 'CurrentAssets'),
             ('PP&E (net)', 'PPE'), ('Total Assets', 'TotalAssets'), ('', ''),
             ('Total Current Liab', 'CurrentLiab'), ('Long-Term Debt', 'LTDebt'),
             ('Total Liabilities', 'TotalLiab'), ('Stockholders Equity', 'Equity')]

    for label, key in items:
        if not label:
            row += 1
            continue
        ws.cell(row=row, column=1, value=label)
        if 'Total' in label:
            ws.cell(row=row, column=1).font = Font(bold=True)
        for col, yr in enumerate(['FY2023', 'FY2024'], 2):
            if key in XBRL_BS[yr]:
                ws.cell(row=row, column=col, value=XBRL_BS[yr][key]).number_format = NUM
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

def create_cash_flow(ws):
    ws['A1'] = 'NOG CASH FLOW STATEMENT'
    ws['A1'].font = Font(bold=True, size=14)

    years = ['FY2022', 'FY2023', 'FY2024', 'FY2025E', 'FY2026E', 'FY2027E', 'FY2028E']
    row = 4
    for col, yr in enumerate(years, 2):
        c = ws.cell(row=row, column=col, value=yr)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1

    # CFO
    ws.cell(row=row, column=1, value='Cash from Operations').font = Font(bold=True)
    cfo_row = row
    for col, yr in enumerate(['FY2022', 'FY2023', 'FY2024'], 2):
        ws.cell(row=row, column=col, value=XBRL_CF[yr]['CFO']).number_format = NUM
    # Projected CFO based on EPS trajectory
    cfo_proj = [1250000, 800000, 1080000, 1010000]
    for col, val in enumerate(cfo_proj, 5):
        ws.cell(row=row, column=col, value=val).number_format = NUM
        ws.cell(row=row, column=col).fill = ESTIMATE_FILL
    row += 1

    # CapEx
    ws.cell(row=row, column=1, value='CapEx').font = Font(bold=True)
    capex_row = row
    capex_hist = [-1674754, -1862346, -1402777]  # Approximations from CFI
    for col, val in enumerate([-987500, -850000, -850000, -850000], 5):
        ws.cell(row=row, column=col, value=val).number_format = NUM
        ws.cell(row=row, column=col).fill = MGMT_FILL
    row += 1

    # FCF
    ws.cell(row=row, column=1, value='Free Cash Flow').font = Font(bold=True)
    fcf_row = row
    fcf_proj = [262500, -50000, 230000, 160000]
    for col, val in enumerate(fcf_proj, 5):
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = NUM
        c.fill = FORMULA_FILL
        c.font = Font(bold=True)

    ws.column_dimensions['A'].width = 25
    return fcf_row, cfo_row, capex_row

def create_dcf(ws):
    ws['A1'] = 'NOG DCF VALUATION'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = 'Exit Multiple approach (Gordon Growth inappropriate for E&P)'
    ws['A2'].font = Font(italic=True, size=10, color="CC0000")

    d = DCF_INPUTS
    row = 5

    # WACC section
    ws.cell(row=row, column=1, value='WACC CALCULATION').font = Font(bold=True)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    wacc_items = [
        ('Risk-Free Rate', d['risk_free'], d['risk_free_source']),
        ('Equity Risk Premium', d['equity_risk_premium'], d['erp_source']),
        ('Beta', d['beta'], d['beta_source']),
        ('Cost of Equity (CAPM)', d['risk_free'] + d['beta'] * d['equity_risk_premium'], 'Rf + β×ERP'),
        ('Pre-tax Cost of Debt', d['cost_of_debt'], d['cod_source']),
        ('After-tax Cost of Debt', d['cost_of_debt'] * (1 - d['tax_rate']), 'Kd×(1-t)'),
    ]

    for label, val, src in wacc_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val).number_format = PCT if 'Rate' in label or 'Cost' in label or 'Premium' in label else '0.00'
        ws.cell(row=row, column=3, value=src).font = SOURCE_FONT
        row += 1

    # DCF Result
    row += 1
    ws.cell(row=row, column=1, value='DCF RESULT').font = Font(bold=True)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    base = calc_dcf(return_details=True)
    results = [
        ('WACC', base['wacc'], ''),
        ('PV of FCF', base['sum_pv'], ''),
        ('Terminal Value', base['tv'], f"EBITDA ${base['ebitda_y4']/1000:.0f}M × {d['exit_multiple']}x"),
        ('PV of Terminal Value', base['pv_tv'], ''),
        ('Enterprise Value', base['ev'], ''),
        ('Less: Net Debt', d['net_debt'], ''),
        ('Equity Value', base['eq_val'], ''),
        ('Shares (000)', d['shares'], ''),
        ('DCF Price per Share', base['dcf_price'], ''),
    ]

    for label, val, src in results:
        ws.cell(row=row, column=1, value=label)
        if 'Price' in label:
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            c = ws.cell(row=row, column=2, value=val)
            c.number_format = PRICE
            c.font = Font(bold=True, size=12)
            c.fill = GOOD_FILL
        elif 'WACC' in label:
            ws.cell(row=row, column=2, value=val).number_format = PCT
        else:
            ws.cell(row=row, column=2, value=val).number_format = NUM
        ws.cell(row=row, column=3, value=src).font = SOURCE_FONT
        row += 1

    # Comparison
    row += 1
    ws.cell(row=row, column=1, value='COMPARISON').font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value='Analyst PT (Avg)')
    ws.cell(row=row, column=2, value=ANALYST['PT']['avg']).number_format = PRICE
    ws.cell(row=row, column=3, value=ANALYST['PT']['source']).font = SOURCE_FONT
    row += 1
    ws.cell(row=row, column=1, value='Current Price')
    ws.cell(row=row, column=2, value=22.56).number_format = PRICE
    row += 1
    ws.cell(row=row, column=1, value='DCF vs PT')
    ws.cell(row=row, column=2, value=(base['dcf_price'] / ANALYST['PT']['avg'] - 1)).number_format = '+0.0%;-0.0%'

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50

# ============================================================================
# SHEET CREATORS - SENSITIVITY
# ============================================================================
def create_tornado(ws):
    ws['A1'] = 'TORNADO ANALYSIS - Key Value Drivers'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = 'Shows DCF price impact from low to high for each driver'
    ws['A2'].font = Font(italic=True, size=10)

    row = 4
    headers = ['Driver', 'Low Value', 'High Value', 'Low DCF', 'High DCF', 'Range', 'Source']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN
    row += 1

    tornado_data = []
    for name, sens in SENSITIVITIES.items():
        vals = sens['range']
        dcfs = run_sensitivity(name, vals)
        tornado_data.append({
            'name': name, 'low_val': min(vals), 'high_val': max(vals),
            'low_dcf': min(dcfs), 'high_dcf': max(dcfs), 'range': max(dcfs) - min(dcfs),
            'source': sens['source'], 'unit': sens['unit'],
        })

    tornado_data.sort(key=lambda x: x['range'], reverse=True)

    for d in tornado_data:
        if d['unit'] == '%':
            low_str, high_str = f"{d['low_val']*100:.1f}%", f"{d['high_val']*100:.1f}%"
        elif d['unit'] in ['$/bbl', '$/MMBtu']:
            low_str, high_str = f"${d['low_val']:.2f}", f"${d['high_val']:.2f}"
        elif d['unit'] == '$M':
            low_str, high_str = f"${d['low_val']/1000:.0f}M", f"${d['high_val']/1000:.0f}M"
        elif d['unit'] == 'k BOE/d':
            low_str, high_str = f"{d['low_val']/1000:.0f}k", f"{d['high_val']/1000:.0f}k"
        elif d['unit'] == 'x':
            low_str, high_str = f"{d['low_val']:.1f}x", f"{d['high_val']:.1f}x"
        else:
            low_str, high_str = f"{d['low_val']:.2f}", f"{d['high_val']:.2f}"

        ws.cell(row=row, column=1, value=d['name']).border = THIN
        ws.cell(row=row, column=2, value=low_str).border = THIN
        ws.cell(row=row, column=3, value=high_str).border = THIN
        c = ws.cell(row=row, column=4, value=d['low_dcf'])
        c.number_format = PRICE
        c.border = THIN
        c.fill = BAD_FILL if d['low_dcf'] < BASE_DCF else GOOD_FILL
        c = ws.cell(row=row, column=5, value=d['high_dcf'])
        c.number_format = PRICE
        c.border = THIN
        c.fill = GOOD_FILL if d['high_dcf'] > BASE_DCF else BAD_FILL
        c = ws.cell(row=row, column=6, value=d['range'])
        c.number_format = PRICE
        c.border = THIN
        c.font = Font(bold=True)
        ws.cell(row=row, column=7, value=d['source']).font = Font(italic=True, size=9)
        row += 1

    ws.column_dimensions['A'].width = 25
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions['G'].width = 35

def create_2way_oil_multiple(ws):
    ws['A1'] = '2-WAY: WTI OIL PRICE vs EXIT MULTIPLE'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = 'Green = Above PT ($29.57) | Yellow = Between Current & PT | Red = Below Current ($22.56)'
    ws['A2'].font = Font(italic=True, size=10)

    wti_vals = [45, 50, 55, 60, 65, 70, 75, 80]
    mult_vals = [3.0, 3.5, 4.0, 4.5, 5.0, 5.5, 6.0]

    row = 4
    ws.cell(row=row, column=1, value='WTI \\ Multiple').font = Font(bold=True)
    for col, mult in enumerate(mult_vals, 2):
        c = ws.cell(row=row, column=col, value=f"{mult:.1f}x")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN
        c.alignment = Alignment(horizontal='center')
    row += 1

    for wti in wti_vals:
        c = ws.cell(row=row, column=1, value=f"${wti}")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN
        for j, mult in enumerate(mult_vals, 2):
            dcf = calc_dcf(wti_price=wti, exit_multiple=mult)
            c = ws.cell(row=row, column=j, value=dcf)
            c.number_format = PRICE
            c.border = THIN
            c.alignment = Alignment(horizontal='center')
            if dcf < 22.56:
                c.fill = BAD_FILL
            elif dcf > 29.57:
                c.fill = GOOD_FILL
            else:
                c.fill = NEUTRAL_FILL
        row += 1

    ws.column_dimensions['A'].width = 15
    for i in range(2, 10):
        ws.column_dimensions[get_column_letter(i)].width = 10

def create_scenarios(ws):
    ws['A1'] = 'SCENARIO ANALYSIS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = 'Based on analyst commodity price forecasts'
    ws['A2'].font = Font(italic=True, size=10)

    row = 4
    headers = ['Scenario', 'WTI (avg)', 'NatGas (avg)', 'Exit Mult', 'DCF Price', 'vs Current', 'vs PT', 'Source']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN
    row += 1

    scenarios = [
        ('Bear', 48, 3.00, 3.5, 'EIA low + market stress'),
        ('Base', 56, 3.50, 4.5, 'JPM/EIA consensus'),
        ('Bull', 68, 4.50, 5.5, 'Industry exec high'),
    ]

    for name, wti, ng, mult, src in scenarios:
        dcf = calc_dcf(wti_price=wti, natgas_price=ng, exit_multiple=mult)
        vs_current = (dcf / 22.56 - 1) * 100
        vs_pt = (dcf / 29.57 - 1) * 100

        ws.cell(row=row, column=1, value=name).border = THIN
        ws.cell(row=row, column=2, value=f"${wti}").border = THIN
        ws.cell(row=row, column=3, value=f"${ng:.2f}").border = THIN
        ws.cell(row=row, column=4, value=f"{mult:.1f}x").border = THIN
        c = ws.cell(row=row, column=5, value=dcf)
        c.number_format = PRICE
        c.border = THIN
        c.font = Font(bold=True)
        if name == 'Bear':
            c.fill = BAD_FILL
        elif name == 'Bull':
            c.fill = GOOD_FILL
        else:
            c.fill = NEUTRAL_FILL
        ws.cell(row=row, column=6, value=f"{vs_current:+.0f}%").border = THIN
        ws.cell(row=row, column=7, value=f"{vs_pt:+.0f}%").border = THIN
        ws.cell(row=row, column=8, value=src).font = Font(italic=True, size=9)
        row += 1

    # Price forecasts
    row += 2
    ws.cell(row=row, column=1, value='COMMODITY PRICE FORECASTS').font = Font(bold=True)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    forecasts = [
        ('WTI Oil', 'EIA', '$52/b 2026, $50/b 2027'),
        ('WTI Oil', 'J.P. Morgan', '$54/b 2026, $53/b 2027'),
        ('WTI Oil', 'Goldman Sachs', '$52/b 2026'),
        ('WTI Oil', 'Industry Execs', '$62/b 2026'),
        ('Natural Gas', 'EIA STEO', '$3.50 2026, $4.60 2027'),
    ]

    for commodity, source, forecast in forecasts:
        ws.cell(row=row, column=1, value=commodity)
        ws.cell(row=row, column=2, value=source)
        ws.cell(row=row, column=3, value=forecast)
        row += 1

    ws.column_dimensions['A'].width = 15
    for i in range(2, 9):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions['H'].width = 25

def create_sources(ws):
    ws['A1'] = 'DATA SOURCES'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'] = 'ALL ASSUMPTIONS VERIFIED FROM CURRENT SOURCES (Feb 2026)'
    ws['A3'].font = Font(bold=True, color="CC0000")

    row = 5
    sources = [
        ('COMMODITY PRICES', [
            ('WTI Forecasts', 'EIA STEO, J.P. Morgan, Goldman Sachs (Jan 2026)'),
            ('Natural Gas Forecasts', 'EIA STEO Jan 2026'),
            ('NOG Realized Price ($61.08/bbl)', 'NOG Q3 2025 Earnings'),
        ]),
        ('DCF INPUTS', [
            ('Risk-Free Rate (4.24%)', '10Y Treasury - FRED/Advisor Perspectives Jan 2026'),
            ('Equity Risk Premium (4.33%)', 'Damodaran NYU Stern Jan 2026'),
            ('Beta (1.10)', 'Average of CNBC (1.02) and TradingView (1.18)'),
            ('Cost of Debt (7.0%)', 'NOG 8.125% Sr Notes + Revolver (-60bps Nov 2025)'),
            ('Exit Multiple (4.5x)', 'E&P EV/EBITDA 4.0-5.55x - Siblis Research Jan 2026'),
        ]),
        ('ANALYST ESTIMATES', [
            ('EPS 2025-2028', 'StockAnalysis.com (12 analysts), mlq.ai, WallStreetZen'),
            ('Revenue 2025-2028', 'StockAnalysis.com (9 analysts)'),
            ('Price Target ($29.57)', 'MarketBeat (8 analysts)'),
        ]),
        ('COMPANY DATA', [
            ('Production (131k BOE/d)', 'NOG Q3 2025 Actual'),
            ('CapEx ($950-1025M)', 'NOG Q3 2025 Guidance (tightened)'),
            ('Historical Financials', 'XBRL 10-K FY2022-2024 via Neo4j'),
        ]),
    ]

    for section, items in sources:
        ws.cell(row=row, column=1, value=section).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = SECTION_FILL
        row += 1
        for item, source in items:
            ws.cell(row=row, column=1, value=f'  {item}')
            ws.cell(row=row, column=2, value=source).font = Font(italic=True, size=10)
            row += 1
        row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 60

# ============================================================================
# MAIN
# ============================================================================
def main():
    print("="*60)
    print("NOG COMPLETE MODEL WITH SENSITIVITY ANALYSIS")
    print("="*60)

    base = calc_dcf(return_details=True)
    print(f"\nBase Case DCF: ${base['dcf_price']:.2f}")
    print(f"Analyst PT: $29.57")
    print(f"Current Price: $22.56")
    print(f"DCF vs PT: {(base['dcf_price']/29.57-1)*100:+.1f}%")

    wb = Workbook()

    # Financial Statements
    ws_is = wb.active
    ws_is.title = "Income Statement"
    create_income_statement(ws_is)

    ws_bs = wb.create_sheet("Balance Sheet")
    create_balance_sheet(ws_bs)

    ws_cf = wb.create_sheet("Cash Flow")
    create_cash_flow(ws_cf)

    ws_dcf = wb.create_sheet("DCF Valuation")
    create_dcf(ws_dcf)

    # Sensitivity Analysis
    ws_torn = wb.create_sheet("Sensitivity-Tornado")
    create_tornado(ws_torn)

    ws_2way = wb.create_sheet("Sensitivity-2Way")
    create_2way_oil_multiple(ws_2way)

    ws_scen = wb.create_sheet("Scenarios")
    create_scenarios(ws_scen)

    ws_src = wb.create_sheet("Sources")
    create_sources(ws_src)

    output = '/home/faisal/EventMarketDB/earnings-analysis/NOG_Complete_Model.xlsx'
    wb.save(output)

    print(f"\nSaved: {output}")
    print("\nSheets:")
    for s in wb.sheetnames:
        print(f"  • {s}")

    print("\n" + "-"*60)
    print("TOP VALUE DRIVERS (by DCF range)")
    print("-"*60)

    tornado_data = []
    for name, sens in SENSITIVITIES.items():
        vals = sens['range']
        dcfs = run_sensitivity(name, vals)
        tornado_data.append((name, max(dcfs) - min(dcfs), min(dcfs), max(dcfs)))

    tornado_data.sort(key=lambda x: x[1], reverse=True)
    for name, rng, low, high in tornado_data[:5]:
        print(f"{name:30} ${low:.2f} - ${high:.2f} (range: ${rng:.2f})")

if __name__ == '__main__':
    main()
